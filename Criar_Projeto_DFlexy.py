#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para gerar SUPER README detalhado do projeto em formato XLSX
Organizado por abas: REDIS, TRACKER, TITLE, DEFINICOES, SCRAPERS
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Cores para formatação
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=8)
TITLE_FONT = Font(bold=True, size=8, color="1F4E78")
SUBHEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=8)
DEFAULT_FONT = Font(size=8)
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_workbook():
    """Cria workbook"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    return wb

def add_title(ws, title, row=1, cols=5):
    """Adiciona título formatado"""
    ws.merge_cells(f'A{row}:{get_column_letter(cols)}{row}')
    cell = ws[f'A{row}']
    cell.value = title
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    return row + 2

def add_subtitle(ws, subtitle, row, cols=5):
    """Adiciona subtítulo formatado"""
    ws.merge_cells(f'A{row}:{get_column_letter(cols)}{row}')
    cell = ws[f'A{row}']
    cell.value = subtitle
    cell.font = SUBHEADER_FONT
    cell.fill = SUBHEADER_FILL
    cell.border = BORDER
    cell.alignment = Alignment(horizontal='left', vertical='center')
    return row + 1

def add_header_row(ws, headers, start_row=1):
    """Adiciona linha de cabeçalho formatada"""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    return start_row + 1

def add_data_row(ws, row_num, data):
    """Adiciona linha de dados formatada"""
    for col_idx, value in enumerate(data, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font = DEFAULT_FONT
        cell.border = BORDER
        cell.alignment = Alignment(vertical='top', wrap_text=True)
    return row_num + 1

def add_tree_text(ws, row, text, indent=0):
    """Adiciona texto de árvore/fluxograma"""
    cell = ws.cell(row=row, column=1, value=text)
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.font = Font(name='Courier New', size=8)
    if indent > 0:
        ws.column_dimensions['A'].width = 100
    return row + 1

def auto_adjust_columns(ws):
    """Ajusta largura das colunas"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 100)
        ws.column_dimensions[column_letter].width = adjusted_width

def apply_default_font(ws):
    """Aplica fonte padrão (tamanho 8) em todas as células que não têm fonte definida"""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and cell.font is None:
                cell.font = DEFAULT_FONT

def create_caracteristicas_sheet(wb):
    """Cria aba Características Principais"""
    ws = wb.create_sheet("Características Principais")
    
    row = add_title(ws, "Características Principais do Projeto", cols=2)
    
    # Lista de características
    row = add_subtitle(ws, "Funcionalidades", row, cols=2)
    headers = ["Característica", "Descrição"]
    row = add_header_row(ws, headers, start_row=row)
    
    caracteristicas_data = [
        ["Múltiplos Scrapers", "Suporte para 7 sites de torrents brasileiros"],
        ["Padronização Inteligente", "Títulos padronizados para facilitar matching automático"],
        ["Metadata API", "Busca automática de tamanhos, datas e nomes via iTorrents.org"],
        ["Tracker Scraping", "Consulta automática de trackers UDP para seeds/leechers"],
        ["FlareSolverr", "Suporte opcional para resolver Cloudflare com sessões reutilizáveis"],
        ["Cache Redis", "Cache inteligente para reduzir carga e latência"],
        ["Circuit Breakers", "Proteção contra sobrecarga de serviços externos"],
        ["Otimizações", "Filtragem antes de enriquecimento pesado para melhor performance"],
    ]
    
    for carac_row in caracteristicas_data:
        row = add_data_row(ws, row, carac_row)
    
    row += 2
    
    # Função Principal
    row = add_subtitle(ws, "Função Principal", row, cols=2)
    headers = ["Componente", "Descrição"]
    row = add_header_row(ws, headers, start_row=row)
    
    funcao_data = [
        ["Função Principal", "Conecta-se a múltiplos sites de torrents e extrai títulos, links magnet, datas, tamanhos e metadados relevantes"],
        ["Trackers", "Consulta trackers UDP automaticamente para preencher seeds/leechers, com cache e lista dinâmica de trackers"],
        ["Padronização", "Padroniza nomes de lançamentos (séries, episódios e filmes) para facilitar matching automático"],
        ["Cache", "Opcionalmente utiliza Redis para cachear o HTML bruto e reduzir carga/latência"],
        ["API", "Expõe uma API JSON simples que pode ser acoplada ao Prowlarr via prowlarr.yml"],
    ]
    
    for funcao_row in funcao_data:
        row = add_data_row(ws, row, funcao_row)
    
    auto_adjust_columns(ws)
    apply_default_font(ws)
    return ws

def create_redis_sheet(wb):
    """Cria aba REDIS - Sistema de Cache"""
    ws = wb.create_sheet("REDIS")
    
    row = add_title(ws, "Sistema de Cache Redis", cols=5)
    
    # Tabela de Cache
    row = add_subtitle(ws, "Tipos de Cache", row, cols=5)
    headers = ["Tipo de Cache", "Chave Redis", "TTL Padrão", "Variável Docker", "Status"]
    row = add_header_row(ws, headers, start_row=row)
    
    cache_data = [
        ["HTML Curto", "html:short:{hash}", "600s (10m)", "HTML_CACHE_TTL_SHORT", "✅ OK"],
        ["HTML Longo", "html:long:{hash}", "43200s (12h)", "HTML_CACHE_TTL_LONG", "✅ OK"],
        ["Metadata (Principal)", "metadata:data:{info_hash}", "604800s (7d)", "Hardcoded", "✅ OK"],
        ["Metadata (Falha)", "metadata:failure:{info_hash}", "60s (1 min)", "Hardcoded", "✅ OK"],
        ["Metadata (Falha 503)", "metadata:failure503:{info_hash}", "300s (5 min)", "Hardcoded", "✅ OK"],
        ["Tracker (Seeds/Leechers)", "tracker:{info_hash}", "604800s (7d)", "Hardcoded", "✅ OK"],
        ["IMDB (por hash)", "imdb:{info_hash}", "604800s (7d)", "Hardcoded", "✅ OK"],
        ["IMDB (por título)", "imdb:title:{hash}", "604800s (7d)", "Hardcoded", "✅ OK"],
        ["Link Protegido", "protlink:{hash}", "604800s (7d)", "Hardcoded", "✅ OK"],
        ["FlareSolverr Session", "flaresolverr:session:{base_url}", "14400s (4h)", "FLARESOLVERR_SESSION_TTL", "✅ OK"],
        ["Tracker List", "tracker:list", "86400s (24h)", "Hardcoded", "✅ OK"],
        ["HTML Failure", "html:failure:{url}", "300s (5 min)", "Hardcoded", "✅ OK"],
        ["Circuit Breaker (Metadata)", "circuit:metadata", "60s (1 min)", "Hardcoded", "✅ OK"],
        ["Circuit Breaker (Tracker)", "circuit:tracker", "60s (1 min)", "Hardcoded", "✅ OK"],
    ]
    
    for cache_row in cache_data:
        row = add_data_row(ws, row, cache_row)
    
    row += 2
    
    # Estratégia de Cache
    row = add_subtitle(ws, "Estratégia de Cache", row, cols=5)
    headers = ["Operação", "Com Redis", "Sem Redis"]
    row = add_header_row(ws, headers, start_row=row)
    
    strategy_data = [
        ["Leitura", "Redis primeiro", "Cache em memória (por requisição)"],
        ["Escrita", "Apenas Redis", "Apenas memória (por requisição)"],
        ["Persistência", "Entre requisições", "Apenas durante requisição atual"],
    ]
    
    for strategy_row in strategy_data:
        row = add_data_row(ws, row, strategy_row)
    
    row += 2
    
    # Comportamento Detalhado por Componente
    row = add_subtitle(ws, "Comportamento Detalhado por Componente", row, cols=5)
    headers = ["Componente", "Leitura (Com Redis)", "Leitura (Sem Redis)", "Escrita (Com Redis)", "Escrita (Sem Redis)"]
    row = add_header_row(ws, headers, start_row=row)
    
    component_behavior_data = [
        ["HTML", "Redis primeiro (longo, depois curto)", "Cache em memória (por requisição)", "Redis (longo ou curto conforme tipo)", "Cache em memória (por requisição)"],
        ["Metadata", "Redis primeiro", "Cache em memória (por requisição)", "Redis (TTL: 7 dias)", "Cache em memória (por requisição)"],
        ["Trackers (Seeds/Leechers)", "Redis primeiro", "Cache em memória (por requisição)", "Redis (TTL: 7 dias)", "Cache em memória (por requisição)"],
        ["Tracker List", "Redis primeiro", "Cache em memória (por requisição)", "Redis (TTL: 24h)", "Cache em memória (por requisição)"],
        ["FlareSolverr Session", "Redis primeiro", "Cache em memória (por requisição)", "Redis (TTL: 4h)", "Cache em memória (por requisição)"],
        ["Link Protegido (ProtLink)", "Redis primeiro", "Cache em memória (por requisição)", "Redis (TTL: 7 dias)", "Cache em memória (por requisição)"],
        ["Circuit Breaker (Metadata)", "Redis primeiro", "Cache em memória (por requisição)", "Redis (TTL: 60s)", "Cache em memória (por requisição)"],
        ["Circuit Breaker (Tracker)", "Redis primeiro", "Cache em memória (por requisição)", "Redis (TTL: 60s)", "Cache em memória (por requisição)"],
        ["Metadata Failure", "Redis primeiro", "Cache em memória (por requisição)", "Redis (TTL: 60s ou 300s)", "Cache em memória (por requisição)"],
        ["HTML Failure", "Redis primeiro", "Cache em memória (por requisição)", "Redis (TTL: 300s)", "Cache em memória (por requisição)"],
    ]
    
    for component_row in component_behavior_data:
        row = add_data_row(ws, row, component_row)
    
    row += 1
    note_cell = ws.cell(row=row, column=1, value="Nota: A memória só é usada se não tiver Redis. Redis sempre tem prioridade. Cache em memória é por requisição (threading.local) e não persiste entre requisições.")
    note_cell.font = Font(italic=True, size=8)
    note_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 2
    
    # Comportamento do Cache de HTML
    row = add_subtitle(ws, "Comportamento do Cache de HTML", row, cols=5)
    headers = ["Situação", "Query", "_is_test", "HTML usa cache?", "Vê novos links?", "Observações"]
    row = add_header_row(ws, headers, start_row=row)
    
    html_behavior_data = [
        ["Busca sem query", "Vazia", "True", "❌ Não (sempre busca fresco)", "✅ Sim", "HTML nunca é salvo no Redis durante buscas sem query"],
        ["Busca com query", "Com query", "False", "✅ Sim (conforme TTL)", "⚠️ Pode demorar (conforme TTL)", "Novos links aparecem quando cache expira"],
    ]
    
    for behavior_row in html_behavior_data:
        row = add_data_row(ws, row, behavior_row)
    
    row += 2
    
    # Exemplo Prático
    row = add_subtitle(ws, "Exemplo Prático (com HTML_CACHE_TTL_LONG=6h)", row, cols=2)
    headers = ["Horário", "Ação"]
    row = add_header_row(ws, headers, start_row=row)
    
    exemplo_data = [
        ["10:00", "Busca com query → Salva cache (válido até 16:00)"],
        ["10:15", "Site adiciona novos links"],
        ["10:30", "Busca com query → Usa cache antigo → ❌ Não vê novos links"],
        ["16:01", "Busca com query → Cache expirou → Busca fresco → ✅ Vê novos links"],
    ]
    
    for exemplo_row in exemplo_data:
        row = add_data_row(ws, row, exemplo_row)
    
    row += 1
    note_cell = ws.cell(row=row, column=1, value="Importante: Durante buscas sem query (`_is_test=True`), o HTML sempre é buscado fresco, garantindo que novos links apareçam imediatamente. O cache de HTML afeta apenas buscas com query.")
    note_cell.font = Font(italic=True, size=8)
    note_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.merge_cells(f'A{row}:B{row}')
    
    row += 2
    
    # Exemplo Prático SHORT
    row = add_subtitle(ws, "Exemplo Prático (com HTML_CACHE_TTL_SHORT=10m)", row, cols=2)
    headers = ["Horário", "Ação"]
    row = add_header_row(ws, headers, start_row=row)
    
    exemplo_short_data = [
        ["10:00", "Busca com query → Salva cache curto (válido até 10:10)"],
        ["10:05", "Site adiciona novos links"],
        ["10:08", "Busca com query → Usa cache curto → ❌ Não vê novos links"],
        ["10:11", "Busca com query → Cache curto expirou → Busca fresco → ✅ Vê novos links"],
    ]
    
    for exemplo_row in exemplo_short_data:
        row = add_data_row(ws, row, exemplo_row)
    
    row += 1
    note_cell = ws.cell(row=row, column=1, value="Nota: O cache curto (SHORT) é usado para páginas de busca e tem TTL menor (padrão 10 minutos). O cache longo (LONG) é usado para páginas de detalhes e tem TTL maior (padrão 12 horas).")
    note_cell.font = Font(italic=True, size=8)
    note_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.merge_cells(f'A{row}:B{row}')
    
    auto_adjust_columns(ws)
    apply_default_font(ws)
    return ws

def create_tracker_sheet(wb):
    """Cria aba TRACKER"""
    ws = wb.create_sheet("TRACKER")
    
    row = add_title(ws, "Sistema de Trackers UDP", cols=4)
    
    # Fontes de Trackers
    row = add_subtitle(ws, "Fontes de Trackers (Ordem de Prioridade)", row, cols=4)
    headers = ["Prioridade", "Fonte", "Método", "Descrição"]
    row = add_header_row(ws, headers, start_row=row)
    
    sources_data = [
        ["1", "Magnet Link", "Parâmetro tr= do magnet", "MagnetParser.parse()"],
        ["2", "Lista Dinâmica", "TrackerListProvider", "Lista de trackers públicos"],
        ["3", "Fallback", "Trackers padrão", "Se nenhum disponível"],
    ]
    
    for source_row in sources_data:
        row = add_data_row(ws, row, source_row)
    
    row += 2
    
    # Processamento
    row = add_subtitle(ws, "Processamento de Trackers", row, cols=4)
    headers = ["Etapa", "Ação", "Observação"]
    row = add_header_row(ws, headers, start_row=row)
    
    process_data = [
        ["1. Extração", "Extrai do magnet ou lista dinâmica", "UDP apenas"],
        ["2. Normalização", "Normaliza URLs", "Remove tokens, padroniza formato"],
        ["3. Busca Peers", "Consulta UDP em lote", "get_peers_bulk()"],
        ["4. Atribuição", "Atribui seeds/leechers", "seed_count, leech_count"],
    ]
    
    for process_row in process_data:
        row = add_data_row(ws, row, process_row)
    
    row += 2
    
    # Circuit Breaker
    row = add_subtitle(ws, "Circuit Breaker - Tracker", row, cols=4)
    headers = ["Tipo", "Threshold", "Duração", "Comportamento"]
    row = add_header_row(ws, headers, start_row=row)
    
    circuit_data = [
        ["Timeouts", "3 consecutivos", "60s", "Usa apenas trackers em cache"],
    ]
    
    for circuit_row in circuit_data:
        row = add_data_row(ws, row, circuit_row)
    
    auto_adjust_columns(ws)
    apply_default_font(ws)
    return ws

def create_title_sheet(wb):
    """Cria aba TITLE - Padronização de Títulos"""
    ws = wb.create_sheet("TITLE")
    
    row = add_title(ws, "Padronização de Títulos", cols=1)
    
    # Fluxograma
    row = add_subtitle(ws, "Fluxograma de Processamento", row, cols=1)
    
    tree = """create_standardized_title()
│
├─ original_title_html existe?
│  │
│  ├─ SIM
│  │  │
│  │  ├─ Tem caracteres não-latinos?
│  │  │  │
│  │  │  ├─ NÃO
│  │  │  │  └─ Usa original_title_html processado
│  │  │  │     ├─ clean_title()
│  │  │  │     ├─ remove_accents()
│  │  │  │     ├─ Remove SxxExx se houver
│  │  │  │     ├─ Remove ano no final
│  │  │  │     ├─ Normaliza (espaços → pontos)
│  │  │  │     └─ Continua processando release_title_magnet
│  │  │  │        └─ Para extrair: SxxExx, ano, informações técnicas
│  │  │  │
│  │  │  └─ SIM
│  │  │     │
│  │  │     ├─ release_title_magnet também tem não-latinos?
│  │  │     │  │
│  │  │     │  ├─ SIM
│  │  │     │  │  │
│  │  │     │  │  ├─ translated_title_html existe?
│  │  │     │  │  │  │
│  │  │     │  │  │  ├─ SIM → Fallback 1.1: Usa translated_title_html processado
│  │  │     │  │  │  │     ├─ clean_title()
│  │  │     │  │  │  │     ├─ remove_accents()
│  │  │     │  │  │  │     ├─ Remove SxxExx se houver
│  │  │     │  │  │  │     ├─ Remove ano no final
│  │  │     │  │  │  │     ├─ Normaliza (espaços → pontos)
│  │  │     │  │  │  │     └─ Continua processando release_title_magnet
│  │  │     │  │  │  │        └─ Para extrair: SxxExx, ano, informações técnicas
│  │  │     │  │  │  │
│  │  │     │  │  │  └─ NÃO → Fallback 1: Usa release_title_magnet
│  │  │     │  │  │        └─ _extract_base_title_from_release()
│  │  │     │  │  │           └─ Continua processando release_title_magnet
│  │  │     │  │  │              └─ Para extrair: SxxExx, ano, informações técnicas
│  │  │     │  │  │
│  │  │     │  │  └─ NÃO → Fallback 1: Usa release_title_magnet
│  │  │     │  │     └─ _extract_base_title_from_release()
│  │  │     │  │        └─ Continua processando release_title_magnet
│  │  │     │  │           └─ Para extrair: SxxExx, ano, informações técnicas
│  │  │     │  │
│  │  │     └─ NÃO → Fallback 1: Usa release_title_magnet
│  │  │        └─ _extract_base_title_from_release()
│  │  │           └─ Continua processando release_title_magnet
│  │  │              └─ Para extrair: SxxExx, ano, informações técnicas
│  │  │
│  │  └─ NÃO → Fallback 2: Usa release_title_magnet
│  │     └─ _extract_base_title_from_release()
│  │        └─ Retorna direto (sem processar mais)
│  │
│  └─ NÃO → Fallback 2: Usa release_title_magnet
│     └─ _extract_base_title_from_release()
│        └─ Retorna direto (sem processar mais)
│
└─ Após determinar base_title (exceto Fallback 2):
   │
   └─ Processa release_title_magnet para extrair:
      ├─ SxxExx (episódios múltiplos ou simples)
      ├─ Sx (temporada completa)
      ├─ Ano (2025, 2024, etc.)
      └─ Informações técnicas (qualidade, codec, fonte, áudio, etc.)
      │
      └─ Monta título final: base_title.SxxExx.ano.qualidade.codec..."""
    
    for line in tree.split('\n'):
        row = add_tree_text(ws, row, line)
        row += 1
    
    row += 2
    
    # Ordem dos Componentes
    row = add_subtitle(ws, "Ordem Final dos Componentes", row, cols=5)
    headers = ["Ordem", "Componente", "Formato", "Exemplo", "Observação"]
    row = add_header_row(ws, headers, start_row=row)
    
    components_data = [
        ["1", "Título Base", "Texto limpo", "OnePunchMan", "Sem acentos, espaços → pontos"],
        ["2", "Temporada/Episódio", "Sxx ou SxxExx", "S03E04", "Zero-padding obrigatório"],
        ["3", "Ano", "19xx ou 20xx", "2025", "Formato 4 dígitos"],
        ["4", "Qualidade", "Minúsculas", "1080p", "Normalizado para minúsculas"],
        ["5", "Fonte", "Padronizado", "WEB-DL", "WEB-DL sempre padronizado"],
        ["6", "Codec", "Minúsculas", "x264", "Normalizado para minúsculas"],
        ["7", "Áudio", "Case original", "LEGENDADO", "Mantém case original"],
        ["8", "Outros", "Variado", "-RARBG", "Release groups e outros"],
        ["9", "Tags de Idioma", "[tag]", "[Eng], [br-dub]", "Adicionado no final"],
    ]
    
    for comp_row in components_data:
        row = add_data_row(ws, row, comp_row)
    
    row += 2
    
    # Exemplos
    row = add_subtitle(ws, "Exemplos de Padronização", row, cols=3)
    headers = ["Tipo", "Entrada Original", "Título Padronizado Final"]
    row = add_header_row(ws, headers, start_row=row)
    
    examples_data = [
        ["Episódio Simples", "original_title='Pluribus', release_title='Pluribus S01E01 WEB-DL 1080p'", "Pluribus.S01E01.2025.1080p.WEB-DL"],
        ["Episódios Múltiplos (2)", "release_title='Pluribus S02E05-06 WEB-DL'", "Pluribus.S02E01-02.2025.WEB-DL"],
        ["Episódios Múltiplos (3-4)", "release_title='Pluribus S02E05-06-07 WEB-DL'", "Pluribus.S02E05E06E07.2025.WEB-DL"],
        ["Episódios Múltiplos (5+)", "release_title='Pluribus S02E01-02-03-04-05 WEB-DL'", "Pluribus.S02E01-E05.2025.WEB-DL"],
        ["Filme", "original_title='Matrix', year='1999', release_title='Matrix 1999 1080p BluRay'", "Matrix.1999.1080p.BluRay"],
        ["Com Não-Latinos", "original_title='ワンパンマン', release_title='OnePunchManS03E04WEB-DL'", "OnePunchMan.S03E04.2025.1080p.WEB-DL"],
        ["Sem Formato", "release_title='Filme.2025'", "Filme.2025.WEB-DL (WEB-DL adicionado)"],
    ]
    
    for example_row in examples_data:
        row = add_data_row(ws, row, example_row)
    
    auto_adjust_columns(ws)
    apply_default_font(ws)
    return ws

def create_definicoes_sheet(wb):
    """Cria aba DEFINICOES - Variáveis e Configurações"""
    ws = wb.create_sheet("DEFINICOES")
    
    row = add_title(ws, "Variáveis de Ambiente e Definições", cols=4)
    
    # Variáveis de Ambiente
    row = add_subtitle(ws, "Variáveis de Ambiente Customizadas", row, cols=4)
    headers = ["Variável", "Descrição", "Padrão", "Localização"]
    row = add_header_row(ws, headers, start_row=row)
    
    variables_data = [
        ["PORT", "Porta do servidor HTTP", "7006", "app/config.py"],
        ["METRICS_PORT", "Porta do servidor de métricas", "8081", "app/config.py"],
        ["REDIS_HOST", "Host do servidor Redis", "localhost", "app/config.py"],
        ["REDIS_PORT", "Porta do servidor Redis", "6379", "app/config.py"],
        ["REDIS_DB", "Database lógica do Redis", "0", "app/config.py"],
        ["HTML_CACHE_TTL_SHORT", "TTL do cache curto de HTML (páginas de busca). Formato: 10m, 1h, 7d", "10m", "app/config.py"],
        ["HTML_CACHE_TTL_LONG", "TTL do cache longo de HTML (páginas de detalhes). Formato: 10m, 1h, 7d", "12h", "app/config.py"],
        ["FLARESOLVERR_SESSION_TTL", "TTL das sessões FlareSolverr. Formato: 10m, 1h, 7d", "4h", "app/config.py"],
        ["EMPTY_QUERY_MAX_LINKS", "Limite de links a processar em query vazia (página 1)", "15", "app/config.py"],
        ["FLARESOLVERR_ADDRESS", "Endereço do servidor FlareSolverr (ex: http://flaresolverr:8191). None = desabilitado", "None (opcional)", "app/config.py"],
        ["LOG_LEVEL", "Nível de log: 0=debug, 1=info, 2=warn, 3=error", "1", "app/config.py"],
        ["LOG_FORMAT", "Formato de log: 'console' ou 'json'", "console", "app/config.py"],
    ]
    
    # Adiciona as variáveis customizadas primeiro
    for var_row in variables_data:
        row = add_data_row(ws, row, var_row)
    
    row += 2
    
    # Valores Hardcoded (não configuráveis)
    row = add_subtitle(ws, "Valores Hardcoded (não configuráveis via Docker)", row, cols=4)
    headers = ["Configuração", "Valor", "Localização", "Descrição"]
    row = add_header_row(ws, headers, start_row=row)
    
    hardcoded_data = [
        ["METADATA_CACHE_TTL", "7 dias (604800s)", "cache/metadata_cache.py", "TTL do cache de metadata"],
        ["TRACKER_CACHE_TTL", "7 dias (604800s)", "cache/tracker_cache.py", "TTL do cache de seeds/leechers"],
        ["IMDB_CACHE_TTL", "7 dias (604800s)", "core/enrichers/torrent_enricher.py", "TTL do cache de IDs IMDB"],
        ["PROTECTED_LINK_CACHE_TTL", "7 dias (604800s)", "utils/parsing/link_resolver.py", "TTL do cache de links protegidos"],
        ["TRACKER_LIST_TTL", "24 horas (86400s)", "tracker/list_provider.py", "TTL da lista dinâmica de trackers"],
        ["HTML_FAILURE_TTL", "5 minutos (300s)", "scraper/base.py", "TTL do cache de falhas ao baixar HTML"],
        ["TRACKER_SCRAPING_ENABLED", "Sempre ON (true)", "scraper/base.py", "Scraping de trackers sempre habilitado"],
    ]
    
    for hardcoded_row in hardcoded_data:
        row = add_data_row(ws, row, hardcoded_row)
    
    row += 2
    
    # Lógicas Principais
    row = add_subtitle(ws, "Lógicas de Processamento", row, cols=3)
    headers = ["Tipo", "Configuração", "Descrição"]
    row = add_header_row(ws, headers, start_row=row)
    
    logic_data = [
        ["Circuit Breaker - Metadata", "Timeouts", "Após 3 timeouts consecutivos ao buscar metadata do iTorrents.org, o sistema desabilita temporariamente a busca de metadata por 60 segundos. Usa Redis para persistir entre requisições ou cache em memória por requisição se Redis não estiver disponível."],
        ["Circuit Breaker - Metadata", "Erros 503", "Após 5 erros 503 consecutivos (Service Unavailable) do iTorrents.org, o sistema desabilita temporariamente a busca de metadata por 60 segundos. Erros 503 são cacheados por 5 minutos no Redis para evitar tentativas repetidas."],
        ["Circuit Breaker - Tracker", "Timeouts", "Após 3 timeouts consecutivos ao buscar lista de trackers remotos, o sistema desabilita temporariamente a busca por 60 segundos. Usa Redis para persistir entre requisições ou cache em memória por requisição se Redis não estiver disponível."],
        ["Rate Limiting - Metadata", "Taxa Base", "Limita requisições ao iTorrents.org para 1 requisição por segundo (intervalo mínimo de 1s entre requisições). Implementado com lock thread-safe para evitar sobrecarga no servidor."],
        ["Rate Limiting - Metadata", "Burst", "Sistema de tokens permite até 2 requisições rápidas consecutivas antes de aplicar o rate limiting. Tokens são recarregados automaticamente (1 token a cada 1 segundo, máximo de 2 tokens)."],
        ["Rate Limiting - Links Protegidos", "Concorrência", "Máximo de 5 requisições simultâneas para resolver links protegidos (systemads, protlink, encurtador). Usa semáforo thread-safe para controlar concorrência."],
        ["Rate Limiting - Links Protegidos", "Delay por Domínio", "Delay mínimo de 200ms entre requisições para o mesmo domínio ao resolver links protegidos. Evita bloqueios por rate limiting dos servidores de encurtamento."],
        ["Processamento Paralelo", "Max Workers (Páginas)", "Processa até 8 páginas HTML simultaneamente usando ThreadPoolExecutor. Cada worker busca e processa uma página independentemente, melhorando performance em buscas com múltiplas páginas."],
        ["Processamento Paralelo", "Max Workers (Metadata)", "Busca metadata (size/date) de até 8 torrents simultaneamente usando ThreadPoolExecutor. Cada worker faz requisição HTTP ao iTorrents.org para obter informações do torrent."],
        ["Processamento Paralelo", "Max Workers (Trackers)", "Faz scraping UDP de até 8 trackers simultaneamente usando ThreadPoolExecutor. Cada worker consulta um tracker UDP para obter seeds/leechers de um torrent."],
        ["Timeout", "Página HTML", "Timeout de 45 segundos para download de páginas HTML dos sites de torrents. Se exceder, a requisição é cancelada e o scraper tenta a próxima página ou retorna erro."],
        ["Timeout", "Metadata", "Timeout de 10 segundos para buscar metadata do iTorrents.org. Se exceder, é registrado como timeout e pode acionar o circuit breaker após 3 ocorrências consecutivas."],
        ["Timeout", "Links Protegidos", "Timeout de 5 segundos por redirect ao resolver links protegidos (aumenta para 10s em links do Twitter/t.co). Máximo de 20 redirects antes de desistir."],
    ]
    
    for logic_row in logic_data:
        row = add_data_row(ws, row, logic_row)
    
    # Ajusta largura da coluna de descrição para textos longos
    ws.column_dimensions['C'].width = 80
    
    row += 2
    
    # Fluxo de Enriquecimento
    row = add_subtitle(ws, "Fluxo de Enriquecimento", row, cols=2)
    headers = ["Etapa", "Ação"]
    row = add_header_row(ws, headers, start_row=row)
    
    enrichment_data = [
        ["1. Remoção Duplicados", "Remove torrents duplicados baseado no info_hash (40 caracteres hexadecimais). Mantém apenas a primeira ocorrência de cada hash único. Torrents sem info_hash válido são mantidos."],
        ["2. Garantia Títulos Completos", "Verifica se cada torrent tem título completo (mínimo 10 caracteres). Se o título estiver vazio ou muito curto, busca metadata do iTorrents.org usando o info_hash para obter o nome completo do torrent. Atualiza o campo 'title' se encontrar um nome válido (mínimo 3 caracteres)."],
        ["3. Aplicação Filtro", "Aplica função de filtro customizada (filter_func) se fornecida. O filtro recebe cada torrent como dicionário e retorna True para manter ou False para remover. Estatísticas de filtragem são armazenadas (total, filtrados, aprovados) para acesso via _last_filter_stats. Se não houver filtro, todos os torrents são aprovados."],
        ["4. Busca Metadata Batch", "Busca metadata (size e date) do iTorrents.org em paralelo usando ThreadPoolExecutor com até 8 workers simultâneos. Processa apenas torrents que ainda não tiveram metadata buscada (_metadata_fetched=False) e que possuem magnet_link. Cada worker faz requisição HTTP ao iTorrents.org com timeout de 10 segundos. Metadata obtida é armazenada em '_metadata' e marcada como '_metadata_fetched=True'."],
        ["5. Aplicação Fallbacks", "Aplica fallbacks em ordem de prioridade para preencher campos ausentes:\n- Size: 1) Metadata API (iTorrents.org - campo 'size'), 2) Parâmetro 'xl' do magnet link, 3) Tamanho extraído do HTML (fallback final)\n- Date: 1) Metadata API (iTorrents.org - campo 'created_time'), 2) Data extraída do HTML (fallback final)\n- IMDB: 1) Cache Redis por info_hash, 2) Cache Redis por base_title normalizado, 3) Metadata do torrent (campo 'imdb' do bencode). Se encontrado, salva no cache Redis por 7 dias para reutilização futura."],
        ["6. Busca Trackers", "Busca seeds/leechers via scraping UDP de trackers em paralelo usando ThreadPoolExecutor com até 8 workers simultâneos. Usa lista dinâmica de trackers (com fallback estático) e filtra apenas trackers UDP válidos. Cada worker consulta múltiplos trackers para um torrent e retorna o melhor resultado (maior número de seeds). Resultados são cacheados no Redis por 7 dias. Seeds e leechers são adicionados aos campos 'seeders' e 'leechers' do torrent."],
    ]
    
    for enrich_row in enrichment_data:
        row = add_data_row(ws, row, enrich_row)
    
    # Ajusta largura da coluna de ação para textos longos
    ws.column_dimensions['B'].width = 100
    
    row += 2
    
    # Busca de Tamanhos
    row = add_subtitle(ws, "Busca de Tamanhos (Ordem de Prioridade)", row, cols=3)
    headers = ["Prioridade", "Fonte", "Método"]
    row = add_header_row(ws, headers, start_row=row)
    
    size_data = [
        ["1", "Metadata API (iTorrents.org)", "Campo 'size' em bytes obtido via fetch_metadata_from_itorrents(). Requisição HTTP ao iTorrents.org com timeout de 10s. Cacheado no Redis por 7 dias."],
        ["2", "Parâmetro 'xl' do Magnet Link", "Parâmetro 'xl' extraído do magnet link via MagnetParser.parse(). Representa o tamanho em bytes do arquivo. Formato: magnet:?xt=urn:btih:...&xl=1234567890"],
        ["3", "Tamanho Extraído do HTML", "Tamanho extraído diretamente do HTML da página pelo scraper. Usado como fallback final se metadata e magnet não tiverem o tamanho."],
    ]
    
    for size_row in size_data:
        row = add_data_row(ws, row, size_row)
    
    row += 2
    
    # Busca de Datas
    row = add_subtitle(ws, "Busca de Datas (Ordem de Prioridade)", row, cols=3)
    headers = ["Prioridade", "Fonte", "Método"]
    row = add_header_row(ws, headers, start_row=row)
    
    date_data = [
        ["1", "Metadata API (iTorrents.org)", "Campo 'created_time' (timestamp Unix) obtido via fetch_metadata_from_itorrents(). Extraído do bencode do torrent (campo 'creation date'). Cacheado no Redis por 7 dias."],
        ["2", "Data Extraída do HTML", "Data extraída diretamente do HTML da página pelo scraper. Pode vir de meta tags, elementos HTML ou texto da página. Usado como fallback final se metadata não tiver a data."],
    ]
    
    for date_row in date_data:
        row = add_data_row(ws, row, date_row)
    
    row += 2
    
    # Busca IMDB - Seção Expandida
    row = add_subtitle(ws, "Busca IMDB - Comportamento Completo", row, cols=4)
    headers = ["Etapa", "Quando Ocorre", "Ação", "Resultado"]
    row = add_header_row(ws, headers, start_row=row)
    
    imdb_comportamento = [
        ["Coleta Inicial", "Durante scraping HTML", "Scraper extrai IMDB do HTML da página (links imdb.com/title/tt1234567). Cada scraper tem sua própria lógica de extração (ex: busca em div.content, próximo a labels 'IMDB', etc.)", "IMDB é adicionado ao torrent['imdb'] se encontrado no HTML"],
        ["Salvamento no Cache", "Se IMDB foi encontrado no HTML", "Salva IMDB em 2 caches Redis:\n1. Por info_hash: 'imdb:{info_hash}' (TTL: 7 dias)\n2. Por base_title: 'imdb:title:{hash_titulo}' (TTL: 7 dias)\nTítulo é normalizado (remove acentos, lowercase, remove componentes técnicos)", "IMDB fica disponível para reutilização em buscas futuras"],
        ["Fallback 1: Cache por info_hash", "Se IMDB não foi encontrado no HTML", "Busca no cache Redis usando chave 'imdb:{info_hash}'. Verifica se existe cache válido (7 dias) para o info_hash do torrent", "Se encontrado: usa IMDB do cache e para busca. Se não encontrado: continua para próximo fallback"],
        ["Fallback 2: Cache por base_title", "Se Fallback 1 não encontrou", "Busca no cache Redis usando chave 'imdb:title:{hash_titulo}'. Título é normalizado (remove acentos, lowercase, remove componentes técnicos como qualidade, codec, etc.) mantendo apenas base_title, temporada/episódio e ano", "Se encontrado: usa IMDB do cache e para busca. Se não encontrado: continua para próximo fallback"],
        ["Fallback 3: Metadata do Torrent", "Se Fallback 2 não encontrou", "Busca IMDB no bencode do torrent via iTorrents.org. Procura por campos: 'imdb', 'imdb_id', 'imdb-id', 'imdb.com'. Usa metadata já buscado se disponível, senão faz requisição HTTP", "Se encontrado: usa IMDB do metadata, salva nos 2 caches Redis (por hash e título) e para busca. Se não encontrado: IMDB permanece vazio"],
        ["Resultado Final", "Após todos os fallbacks", "Se nenhum fallback encontrou IMDB, o campo 'imdb' permanece vazio (string vazia ou ausente) no torrent retornado", "Torrent sem IMDB é retornado normalmente (IMDB é opcional)"],
    ]
    
    for imdb_row in imdb_comportamento:
        row = add_data_row(ws, row, imdb_row)
    
    # Ajusta largura das colunas para textos longos
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 60
    
    row += 2
    
    # Ordem de Prioridade (Resumo)
    row = add_subtitle(ws, "Ordem de Prioridade - Resumo", row, cols=3)
    headers = ["Prioridade", "Fonte", "Método"]
    row = add_header_row(ws, headers, start_row=row)
    
    imdb_prioridade = [
        ["1", "HTML da Página (Scraper)", "ID IMDB extraído diretamente do HTML durante scraping. Cada scraper implementa sua própria lógica de busca (ex: bludv busca em div.content, tfilme busca próximo a <strong>IMDb</strong>, etc.)"],
        ["2", "Cache Redis por info_hash", "Busca no cache usando chave 'imdb:{info_hash}'. Cache válido por 7 dias. Reutiliza IMDB de torrents anteriores com mesmo info_hash."],
        ["3", "Cache Redis por base_title", "Busca no cache usando chave 'imdb:title:{hash_titulo}'. Título normalizado (remove acentos, lowercase, remove componentes técnicos). Permite reutilizar IMDB de torrents com mesmo título base mesmo com info_hash diferente."],
        ["4", "Metadata do Torrent (bencode)", "ID IMDB extraído de campos customizados no bencode do torrent via iTorrents.org. Busca por campos: 'imdb', 'imdb_id', 'imdb-id', 'imdb.com'. Se encontrado, salva nos 2 caches Redis para reutilização futura."],
    ]
    
    for imdb_row in imdb_prioridade:
        row = add_data_row(ws, row, imdb_row)
    
    # Ajusta largura das colunas
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 80
    
    row += 2
    
    # Quando IMDB NÃO é encontrado
    row = add_subtitle(ws, "Quando IMDB NÃO é Encontrado", row, cols=2)
    headers = ["Situação", "Comportamento"]
    row = add_header_row(ws, headers, start_row=row)
    
    imdb_nao_encontrado = [
        ["HTML não contém link IMDB", "Scraper não encontra links imdb.com/title/tt1234567 no HTML da página. Sistema tenta fallbacks (cache e metadata)"],
        ["Cache por info_hash vazio", "Nenhum torrent anterior com mesmo info_hash tinha IMDB cacheado. Sistema tenta próximo fallback (cache por título)"],
        ["Cache por base_title vazio", "Nenhum torrent anterior com mesmo título base tinha IMDB cacheado. Sistema tenta último fallback (metadata)"],
        ["Metadata não contém IMDB", "Bencode do torrent não possui campos IMDB customizados. Sistema finaliza busca e retorna torrent sem IMDB"],
        ["Resultado Final", "Campo 'imdb' permanece vazio (string vazia ou ausente) no torrent. Torrent é retornado normalmente - IMDB é opcional e não impede o funcionamento"],
    ]
    
    for imdb_row in imdb_nao_encontrado:
        row = add_data_row(ws, row, imdb_row)
    
    row += 2
    
    # Normalização do Título para Cache
    row = add_subtitle(ws, "Normalização do Título para localizar o IMDB em Cache (base_title)", row, cols=2)
    headers = ["Operação", "Exemplo"]
    row = add_header_row(ws, headers, start_row=row)
    
    imdb_normalizacao = [
        ["Título Original", "Pluribus.S01E01.2025.1080p.WEB-DL.x264.LEGENDADO"],
        ["Remove tags de áudio", "Pluribus.S01E01.2025.1080p.WEB-DL.x264"],
        ["Remove componentes técnicos", "Pluribus.S01E01.2025"],
        ["Remove acentos", "Pluribus.S01E01.2025 (mantém mesmo se tiver acentos)"],
        ["Converte para lowercase", "pluribus.s01e01.2025"],
        ["Resultado Final (base_title)", "pluribus.s01e01.2025"],
        ["Uso no Cache", "Chave: 'imdb:title:{hash_do_base_title}' - Permite reutilizar IMDB de qualquer torrent com mesmo título base, temporada/episódio e ano"],
    ]
    
    for imdb_row in imdb_normalizacao:
        row = add_data_row(ws, row, imdb_row)
    
    auto_adjust_columns(ws)
    apply_default_font(ws)
    return ws

def create_scrapers_sheet(wb):
    """Cria aba SCRAPERS"""
    ws = wb.create_sheet("SCRAPERS")
    
    row = add_title(ws, "Scrapers Suportados", cols=4)
    
    # Tabela de Scrapers
    row = add_subtitle(ws, "Lista de Scrapers", row, cols=4)
    headers = ["ID", "Nome", "Domínio", "Observações"]
    row = add_header_row(ws, headers, start_row=row)
    
    scrapers_data = [
        ["1", "Starck Filmes", "https://starckfilmes-v3.com/", "Magnet direto no HTML, títulos completos"],
        ["2", "Rede Torrent", "https://redetorrent.com/", "Estrutura padrão"],
        ["3", "Baixa Filmes", "https://www.baixafilmestorrent.com.br/", "Links via Vialink (token base64), temporada no HTML"],
        ["4", "Torrent dos Filmes", "https://torrentdosfilmes.se/", "Precisa tratar metas de temporada/ano"],
        ["6", "Comando Torrents", "https://comando.la/", "Múltiplos magnets por página, título original em HTML"],
        ["7", "Bludv Filmes", "https://bludv.net/", "Links protegidos systemads, magnet em data-download"],
    ]
    
    for scraper_row in scrapers_data:
        row = add_data_row(ws, row, scraper_row)
    
    # Adiciona informações sobre SCRAPER_TYPE de cada scraper
    row += 1
    note_cell = ws.cell(row=row, column=1, value="Nota: Os IDs numéricos (1-7) são mapeados para SCRAPER_TYPE no arquivo api/services/indexer_service.py (SCRAPER_NUMBER_MAP)")
    note_cell.font = Font(italic=True, size=9, color="0066CC")
    note_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 2
    
    # Fluxo de Busca
    row = add_subtitle(ws, "Fluxo de Busca por Query", row, cols=2)
    headers = ["Etapa", "Ação"]
    row = add_header_row(ws, headers, start_row=row)
    
    flow_data = [
        ["1", "Request HTTP → api/handlers.py"],
        ["2", "Handler → api/services/indexer_service.py"],
        ["3", "Service → scraper/{site}.py (search)"],
        ["4", "Scraper → Extrai HTML → Parsing → Lista de torrents brutos"],
        ["5", "Scraper → core/enrichers/torrent_enricher.py"],
        ["6", "Enricher → Remove duplicados → Busca títulos → Filtra → Busca metadata → Busca trackers"],
        ["7", "Service → core/processors/torrent_processor.py"],
        ["8", "Processor → Remove campos internos → Ordena por data"],
        ["9", "Handler → Response JSON"],
    ]
    
    for flow_row in flow_data:
        row = add_data_row(ws, row, flow_row)
    
    row += 2
    
    # Fluxo de Query Vazia
    row = add_subtitle(ws, "Fluxo de Lista de Página (Query Vazia)", row, cols=2)
    headers = ["Etapa", "Ação"]
    row = add_header_row(ws, headers, start_row=row)
    
    test_flow_data = [
        ["1", "Request HTTP → api/handlers.py"],
        ["2", "Handler → api/services/indexer_service.py"],
        ["3", "Service → scraper/{site}.py (get_page)"],
        ["4", "Scraper → Extrai HTML → Parsing → Lista de links"],
        ["5", "Scraper → utils/concurrency/scraper_helpers.py (limita links)"],
        ["6", "Scraper → Processa links sequencialmente (mantém ordem)"],
        ["7", "Scraper → core/enrichers/torrent_enricher.py (skip_metadata=True)"],
        ["8", "Service → Mantém ordem original (não ordena)"],
        ["9", "Handler → Response JSON"],
    ]
    
    for test_flow_row in test_flow_data:
        row = add_data_row(ws, row, test_flow_row)
    
    row += 2
    
    # API Endpoints
    row = add_subtitle(ws, "API Endpoints", row, cols=3)
    headers = ["Método", "Rota", "Descrição"]
    row = add_header_row(ws, headers, start_row=row)
    
    api_data = [
        ["GET", "/", "Informações básicas da API"],
        ["GET", "/indexer", "Usa scraper padrão"],
        ["GET", "/indexer?q=foo", "Busca na fonte padrão"],
        ["GET", "/indexer?page=2", "Paginação"],
        ["GET", "/indexer?q=foo&filter_results=true", "Busca com filtro"],
        ["GET", "/indexer?q=foo&use_flaresolverr=true", "Busca com FlareSolverr"],
        ["GET", "/indexers/<tipo>?q=foo", "Usa scraper específico (1-7)"],
    ]
    
    for api_row in api_data:
        row = add_data_row(ws, row, api_row)
    
    row += 2
    
    # Filtro de Resultados
    row = add_subtitle(ws, "Filtro de Resultados", row, cols=3)
    headers = ["Quantidade de Palavras", "Regra", "Exemplo"]
    row = add_header_row(ws, headers, start_row=row)
    
    filter_data = [
        ["1 palavra", "Exige que corresponda", "pluribus → deve estar no título"],
        ["2 palavras", "Exige que ambas correspondam", "pluribus temporada → ambas devem estar"],
        ["3+ palavras", "Exige que pelo menos 2 correspondam", "pluribus temporada 1 → pelo menos 2 devem estar"],
    ]
    
    for filter_row in filter_data:
        row = add_data_row(ws, row, filter_row)
    
    row += 2
    
    # Sistema de IDs dos Scrapers
    row = add_subtitle(ws, "Sistema de IDs dos Scrapers", row, cols=4)
    headers = ["Componente", "Localização", "Descrição", "Exemplo"]
    row = add_header_row(ws, headers, start_row=row)
    
    id_system_data = [
        ["SCRAPER_NUMBER_MAP", "api/services/indexer_service.py", "Mapeamento de IDs numéricos (1-7) para SCRAPER_TYPE. Usado pelo Prowlarr para identificar scrapers via URL /indexers/<id>", '"1": "starck", "2": "rede", ...'],
        ["SCRAPER_TYPE", "scraper/{site}.py", "Atributo de classe que define o identificador único do scraper. Deve ser único e em minúsculas", 'SCRAPER_TYPE = "starck"'],
        ["Descoberta Automática", "scraper/__init__.py", "Sistema descobre automaticamente todos os scrapers no diretório scraper/ que herdam de BaseScraper", "_discover_scrapers()"],
        ["Normalização", "scraper/__init__.py", "Nomes são normalizados: minúsculas, sem hífens, espaços substituídos por underscore", '"Starck-Filmes" → "starck_filmes"'],
    ]
    
    for id_row in id_system_data:
        row = add_data_row(ws, row, id_row)
    
    row += 2
    
    # Mapeamento Detalhado ID → SCRAPER_TYPE
    row = add_subtitle(ws, "Mapeamento Detalhado: ID → SCRAPER_TYPE → Classe", row, cols=5)
    headers = ["ID Numérico", "SCRAPER_TYPE", "Nome da Classe", "Módulo", "DISPLAY_NAME"]
    row = add_header_row(ws, headers, start_row=row)
    
    mapping_data = [
        ["1", "starck", "StarckScraper", "scraper/starck.py", "Starck"],
        ["2", "rede", "RedeScraper", "scraper/rede.py", "Rede"],
        ["3", "baixafilmes", "BaixafilmesScraper", "scraper/baixafilmes.py", "Baixa Filmes"],
        ["4", "tfilme", "TfilmeScraper", "scraper/tfilme.py", "TFilme"],
        ["6", "comand", "ComandScraper", "scraper/comand.py", "Comando"],
        ["7", "bludv", "BludvScraper", "scraper/bludv.py", "Bludv"],
    ]
    
    for map_row in mapping_data:
        row = add_data_row(ws, row, map_row)
    
    row += 2
    
    # Fluxo Completo para Inserir Novo Scraper
    row = add_subtitle(ws, "FLUXO COMPLETO: Como Inserir um Novo Scraper", row, cols=4)
    headers = ["#", "Passo", "Arquivo/Ação", "Detalhes e Exemplo"]
    row = add_header_row(ws, headers, start_row=row)
    
    fluxo_data = [
        ["1", "Analisar o site de torrents", "Navegador/Inspecionar", "Acesse o site e inspecione o HTML:\n- Como funciona a busca? (URL, parâmetros)\n- Onde estão os links dos torrents?\n- Onde está o magnet link?\n- Qual a estrutura HTML das páginas?\n- Precisa de FlareSolverr (Cloudflare)?"],
        ["2", "Criar arquivo do scraper", "scraper/novosite.py", "Criar novo arquivo Python no diretório scraper/\nNome do arquivo será usado como fallback se SCRAPER_TYPE não for definido\nExemplo: scraper/meusite.py"],
        ["3", "Definir classe do scraper", "scraper/novosite.py", "class MeuSiteScraper(BaseScraper):\n    SCRAPER_TYPE = \"meusite\"\n    DEFAULT_BASE_URL = \"https://meusite.com/\"\n    DISPLAY_NAME = \"Meu Site\""],
        ["4", "Implementar método search()", "scraper/novosite.py", "def search(self, query: str, filter_func=None):\n    # Opção 1: Usar implementação padrão\n    return self._default_search(query, filter_func)\n    # Opção 2: Implementação customizada\n    # ... sua lógica aqui"],
        ["5", "Implementar método get_page()", "scraper/novosite.py", "def get_page(self, page='1', max_items=None):\n    # Opção 1: Usar implementação padrão\n    return self._default_get_page(page, max_items)\n    # Opção 2: Implementação customizada"],
        ["6", "Implementar _extract_links_from_page()", "scraper/novosite.py", "def _extract_links_from_page(self, doc):\n    links = []\n    for item in doc.select('.torrent-item'):\n        link = item.select_one('a[href*=\"/torrent/\"]')\n        if link:\n            links.append(link.get('href'))\n    return links"],
        ["7", "Implementar _parse_torrent_page()", "scraper/novosite.py", "def _parse_torrent_page(self, url):\n    doc = self.get_document(url, self.base_url)\n    if not doc:\n        return None\n    \n    title = doc.select_one('h1').get_text(strip=True)\n    magnet = doc.select_one('a[href^=\"magnet:\"]').get('href')\n    \n    return {\n        'title': title,\n        'magnet_link': magnet,\n        'url': url\n    }"],
        ["8", "(Opcional) Adicionar ID numérico", "api/services/indexer_service.py", "Adicionar no SCRAPER_NUMBER_MAP:\nSCRAPER_NUMBER_MAP[\"8\"] = \"meusite\"\n\nIsso permite usar /indexers/8 na API\nSe não adicionar, use /indexers/meusite"],
        ["9", "Testar descoberta automática", "Logs do servidor", "Reinicie o servidor e verifique logs:\n'Scrapers disponíveis: [..., meusite, ...]'\n\nO sistema descobre automaticamente via scraper/__init__.py"],
        ["10", "Testar busca por nome", "API/Postman", "GET /indexers/meusite?q=teste\n\nDeve retornar resultados ou lista vazia\nVerificar logs para erros"],
        ["11", "Testar busca por ID (se adicionou)", "API/Postman", "GET /indexers/8?q=teste\n\nDeve funcionar se adicionou ao SCRAPER_NUMBER_MAP"],
        ["12", "Testar query vazia (Prowlarr)", "API/Postman", "GET /indexers/meusite\n\nDeve retornar lista de torrents da primeira página\nUsado pelo Prowlarr para testar conexão"],
        ["13", "Verificar enriquecimento", "Logs/Resposta JSON", "Verificar se campos estão sendo preenchidos:\n- title, magnet_link, url (obrigatórios)\n- size, date (via metadata API)\n- seeders, leechers (via tracker scraping)\n- info_hash (extraído do magnet)"],
        ["14", "Testar com FlareSolverr (se necessário)", "API/Postman", "GET /indexers/meusite?q=teste&use_flaresolverr=true\n\nSe o site usa Cloudflare, configure FLARESOLVERR_ADDRESS\nno docker-compose.yml ou variáveis de ambiente"],
    ]
    
    for fluxo_row in fluxo_data:
        row = add_data_row(ws, row, fluxo_row)
    
    # Ajusta largura das colunas para textos longos
    ws.column_dimensions['D'].width = 100
    
    row += 2
    
    # Checklist de Verificação
    row = add_subtitle(ws, "Checklist de Verificação do Novo Scraper", row, cols=3)
    headers = ["Item", "Status", "Observações"]
    row = add_header_row(ws, headers, start_row=row)
    
    checklist_data = [
        ["✅ SCRAPER_TYPE definido e único", "Obrigatório", "Deve ser minúsculas, sem espaços/hífens. Verificar que não existe outro scraper com mesmo SCRAPER_TYPE"],
        ["✅ DEFAULT_BASE_URL válida", "Obrigatório", "URL deve terminar com / e ser acessível"],
        ["✅ DISPLAY_NAME definido", "Obrigatório", "Nome amigável para exibição"],
        ["✅ Classe herda de BaseScraper", "Obrigatório", "class MeuScraper(BaseScraper):"],
        ["✅ Método search() implementado", "Obrigatório", "Pode usar _default_search() ou implementação customizada"],
        ["✅ Método get_page() implementado", "Obrigatório", "Pode usar _default_get_page() ou implementação customizada"],
        ["✅ _extract_links_from_page() implementado", "Se usar _default_get_page()", "Extrai links da página inicial"],
        ["✅ _parse_torrent_page() implementado", "Se usar _default_search()", "Extrai dados do torrent da página"],
        ["✅ Tratamento de erros", "Recomendado", "Timeouts, HTML inválido, elementos não encontrados"],
        ["✅ Testado busca com query", "Obrigatório", "GET /indexers/meusite?q=teste retorna resultados"],
        ["✅ Testado query vazia", "Obrigatório", "GET /indexers/meusite retorna lista de página"],
        ["✅ Verificado enriquecimento", "Recomendado", "Campos size, date, seeders, leechers sendo preenchidos"],
        ["✅ ID numérico adicionado (opcional)", "Opcional", "Se quiser usar /indexers/8, adicionar ao SCRAPER_NUMBER_MAP"],
        ["✅ FlareSolverr configurado (se necessário)", "Se site usa Cloudflare", "Configurar FLARESOLVERR_ADDRESS e testar"],
    ]
    
    for check_row in checklist_data:
        row = add_data_row(ws, row, check_row)
    
    row += 2
    
    # Arquivos Modificados/Criados
    row = add_subtitle(ws, "Arquivos que Precisam ser Criados/Modificados", row, cols=3)
    headers = ["Arquivo", "Ação", "O que Fazer"]
    row = add_header_row(ws, headers, start_row=row)
    
    arquivos_data = [
        ["scraper/novosite.py", "CRIAR", "Criar novo arquivo com a classe do scraper. Nome do arquivo pode ser qualquer coisa, mas SCRAPER_TYPE deve ser único"],
        ["api/services/indexer_service.py", "MODIFICAR (opcional)", "Adicionar entrada no SCRAPER_NUMBER_MAP se quiser usar ID numérico:\nSCRAPER_NUMBER_MAP[\"8\"] = \"novosite\""],
        ["scraper/__init__.py", "NENHUMA", "Não precisa modificar! O sistema descobre automaticamente todos os scrapers que herdam de BaseScraper"],
        ["api/handlers.py", "NENHUMA", "Não precisa modificar! Funciona automaticamente com qualquer scraper descoberto"],
        ["docker-compose.yml", "MODIFICAR (se necessário)", "Apenas se precisar configurar FlareSolverr ou outras variáveis de ambiente específicas"],
    ]
    
    for arquivo_row in arquivos_data:
        row = add_data_row(ws, row, arquivo_row)
    
    # Ajusta largura das colunas
    ws.column_dimensions['C'].width = 80
    
    row += 2
    
    # Como Criar um Novo Scraper
    row = add_subtitle(ws, "Como Criar um Novo Scraper (Detalhes Técnicos)", row, cols=3)
    headers = ["Passo", "Ação", "Detalhes"]
    row = add_header_row(ws, headers, start_row=row)
    
    create_scraper_data = [
        ["1", "Criar arquivo Python", "Criar arquivo scraper/{nome_site}.py (ex: scraper/novosite.py). Nome do arquivo será usado como fallback se SCRAPER_TYPE não for definido"],
        ["2", "Criar classe que herda BaseScraper", "class NovoScraper(BaseScraper): ... A classe deve herdar de BaseScraper e implementar métodos necessários"],
        ["3", "Definir SCRAPER_TYPE", "SCRAPER_TYPE = \"novosite\" (deve ser único, minúsculas, sem espaços/hífens). Este é o identificador usado na API"],
        ["4", "Definir DEFAULT_BASE_URL", "DEFAULT_BASE_URL = \"https://exemplo.com/\" (URL base do site de torrents, deve terminar com /)"],
        ["5", "Definir DISPLAY_NAME", "DISPLAY_NAME = \"Novo Site\" (nome amigável para exibição na API e logs)"],
        ["6", "Implementar métodos obrigatórios", "Implementar search() e/ou get_page() conforme necessário. Ver scraper/base.py para métodos auxiliares disponíveis"],
        ["7", "Implementar métodos de parsing", "Implementar _extract_links_from_page(), _parse_torrent_page(), etc. conforme estrutura HTML do site"],
        ["8", "(Opcional) Adicionar ao SCRAPER_NUMBER_MAP", "Se quiser usar ID numérico (ex: /indexers/8), adicionar em api/services/indexer_service.py: SCRAPER_NUMBER_MAP[\"8\"] = \"novosite\""],
        ["9", "Testar descoberta automática", "O sistema descobre automaticamente o scraper ao iniciar. Verificar logs: 'Scrapers disponíveis: ...'"],
    ]
    
    for create_row in create_scraper_data:
        row = add_data_row(ws, row, create_row)
    
    row += 2
    
    # Estrutura Mínima de um Scraper
    row = add_subtitle(ws, "Estrutura Mínima de um Scraper", row, cols=2)
    headers = ["Componente", "Código Exemplo"]
    row = add_header_row(ws, headers, start_row=row)
    
    structure_data = [
        ["Imports básicos", "from scraper.base import BaseScraper\nfrom typing import List, Dict, Optional, Callable\nfrom bs4 import BeautifulSoup"],
        ["Definição da classe", "class NovoScraper(BaseScraper):\n    SCRAPER_TYPE = \"novosite\"\n    DEFAULT_BASE_URL = \"https://exemplo.com/\"\n    DISPLAY_NAME = \"Novo Site\""],
        ["Método search()", "def search(self, query: str, filter_func: Optional[Callable] = None) -> List[Dict]:\n    return self._default_search(query, filter_func)"],
        ["Método get_page()", "def get_page(self, page: str = '1', max_items: Optional[int] = None) -> List[Dict]:\n    return self._default_get_page(page, max_items)"],
        ["Extração de links", "def _extract_links_from_page(self, doc: BeautifulSoup) -> List[str]:\n    # Implementar lógica para extrair links da página\n    return links"],
        ["Parsing de torrent", "def _parse_torrent_page(self, url: str) -> Optional[Dict]:\n    # Implementar lógica para extrair dados do torrent\n    return torrent_dict"],
    ]
    
    for struct_row in structure_data:
        row = add_data_row(ws, row, struct_row)
    
    # Ajusta largura das colunas para código
    ws.column_dimensions['B'].width = 80
    
    row += 2
    
    # Métodos Auxiliares Disponíveis
    row = add_subtitle(ws, "Métodos Auxiliares Disponíveis (BaseScraper)", row, cols=3)
    headers = ["Método", "Descrição", "Uso"]
    row = add_header_row(ws, headers, start_row=row)
    
    helper_methods_data = [
        ["get_document(url)", "Obtém HTML do cache ou faz requisição HTTP. Suporta FlareSolverr e cache Redis", "doc = self.get_document(url, referer)"],
        ["_default_search(query)", "Implementação padrão de busca com variações de query e paginação", "return self._default_search(query, filter_func)"],
        ["_default_get_page(page)", "Implementação padrão de get_page com extração de links e parsing", "return self._default_get_page(page, max_items)"],
        ["_search_variations(query)", "Gera variações da query (remove stop words, primeira palavra, etc.)", "variations = self._search_variations(query)"],
        ["_extract_links_from_page(doc)", "Método abstrato: deve ser implementado para extrair links da página", "links = self._extract_links_from_page(doc)"],
        ["_parse_torrent_page(url)", "Método abstrato: deve ser implementado para extrair dados do torrent", "torrent = self._parse_torrent_page(url)"],
        ["_enrich_torrents(torrents)", "Enriquece lista de torrents com metadata, trackers, etc.", "self._enrich_torrents(torrents)"],
    ]
    
    for helper_row in helper_methods_data:
        row = add_data_row(ws, row, helper_row)
    
    row += 2
    
    # Requisitos para Novo Scraper
    row = add_subtitle(ws, "Requisitos e Boas Práticas", row, cols=2)
    headers = ["Requisito", "Descrição"]
    row = add_header_row(ws, headers, start_row=row)
    
    requirements_data = [
        ["SCRAPER_TYPE único", "Cada scraper deve ter um SCRAPER_TYPE único. Não pode haver duplicatas"],
        ["DEFAULT_BASE_URL válida", "URL deve ser válida e terminar com /. Será usada como base para todas as requisições"],
        ["Herança de BaseScraper", "Classe deve herdar de BaseScraper para ter acesso a métodos auxiliares e cache"],
        ["Implementar métodos de parsing", "Pelo menos _extract_links_from_page() ou _parse_torrent_page() devem ser implementados"],
        ["Tratamento de erros", "Implementar tratamento de erros adequado (timeouts, HTML inválido, etc.)"],
        ["Suporte a FlareSolverr", "Se o site usa Cloudflare, o scraper automaticamente suporta FlareSolverr se configurado"],
        ["Cache automático", "Cache de HTML é automático via Redis (se disponível). Não precisa implementar manualmente"],
        ["Enriquecimento automático", "Torrents são automaticamente enriquecidos com metadata, trackers, etc. via TorrentEnricher"],
    ]
    
    for req_row in requirements_data:
        row = add_data_row(ws, row, req_row)
    
    row += 2
    
    # Exemplos Práticos de Implementação
    row = add_subtitle(ws, "Exemplos Práticos de Implementação", row, cols=2)
    headers = ["Cenário", "Código Exemplo"]
    row = add_header_row(ws, headers, start_row=row)
    
    exemplos_praticos = [
        ["Site com busca simples (?s=query)", "def __init__(self, base_url=None, use_flaresolverr=False):\n    super().__init__(base_url, use_flaresolverr)\n    self.search_url = \"?s=\"\n    self.page_pattern = \"page/{}/\"\n\ndef search(self, query, filter_func=None):\n    return self._default_search(query, filter_func)"],
        ["Site com busca em /search/query", "def __init__(self, base_url=None, use_flaresolverr=False):\n    super().__init__(base_url, use_flaresolverr)\n    self.search_url = \"search/\"\n    self.page_pattern = \"search/{}/\"\n\ndef search(self, query, filter_func=None):\n    return self._default_search(query, filter_func)"],
        ["Extrair links de lista com classe .item", "def _extract_links_from_page(self, doc):\n    links = []\n    for item in doc.select('.item'):\n        link = item.select_one('a[href*=\"/torrent/\"]')\n        if link:\n            href = link.get('href')\n            if href:\n                links.append(href)\n    return links"],
        ["Extrair links de div com data-id", "def _extract_links_from_page(self, doc):\n    links = []\n    for div in doc.select('div[data-id]'):\n        link = div.select_one('a')\n        if link and link.get('href'):\n            links.append(link.get('href'))\n    return links"],
        ["Extrair magnet de link direto", "def _parse_torrent_page(self, url):\n    doc = self.get_document(url, self.base_url)\n    if not doc:\n        return None\n    \n    magnet = doc.select_one('a[href^=\"magnet:\"]')\n    if not magnet:\n        return None\n    \n    return {'magnet_link': magnet.get('href'), 'url': url}"],
        ["Extrair magnet de atributo data-download", "def _parse_torrent_page(self, url):\n    doc = self.get_document(url, self.base_url)\n    if not doc:\n        return None\n    \n    button = doc.select_one('button[data-download]')\n    if button:\n        magnet = button.get('data-download')\n        return {'magnet_link': magnet, 'url': url}\n    return None"],
        ["Extrair título e magnet juntos", "def _parse_torrent_page(self, url):\n    doc = self.get_document(url, self.base_url)\n    if not doc:\n        return None\n    \n    title = doc.select_one('h1').get_text(strip=True)\n    magnet = doc.select_one('a[href^=\"magnet:\"]').get('href')\n    \n    return {\n        'title': title,\n        'magnet_link': magnet,\n        'url': url\n    }"],
        ["Tratamento de erros robusto", "def _parse_torrent_page(self, url):\n    try:\n        doc = self.get_document(url, self.base_url)\n        if not doc:\n            return None\n        \n        title_elem = doc.select_one('h1')\n        magnet_elem = doc.select_one('a[href^=\"magnet:\"]')\n        \n        if not title_elem or not magnet_elem:\n            return None\n        \n        return {\n            'title': title_elem.get_text(strip=True),\n            'magnet_link': magnet_elem.get('href'),\n            'url': url\n        }\n    except Exception as e:\n        logger.error(f\"Erro ao parsear {url}: {e}\")\n        return None"],
    ]
    
    for exemplo_row in exemplos_praticos:
        row = add_data_row(ws, row, exemplo_row)
    
    # Ajusta largura das colunas para código
    ws.column_dimensions['B'].width = 100
    
    auto_adjust_columns(ws)
    apply_default_font(ws)
    return ws

def create_prowlarr_sheet(wb):
    """Cria aba PROWLARR - Integração com Prowlarr"""
    ws = wb.create_sheet("PROWLARR")
    
    row = add_title(ws, "Integração com Prowlarr", cols=4)
    
    # O que é Prowlarr
    row = add_subtitle(ws, "O que é Prowlarr?", row, cols=4)
    headers = ["Conceito", "Descrição"]
    row = add_header_row(ws, headers, start_row=row)
    
    prowlarr_info = [
        ["Definição", "Prowlarr é um gerenciador de indexadores de torrents que centraliza a busca de conteúdo para aplicações como Sonarr (séries) e Radarr (filmes)"],
        ["Função", "O Prowlarr se conecta a múltiplos indexadores (incluindo este projeto) e permite que Sonarr/Radarr busquem torrents de forma unificada"],
        ["Arquivo de Configuração", "O arquivo prowlarr.yml define como o Prowlarr deve se conectar à API deste projeto e como interpretar os resultados"],
        ["Localização", "O arquivo prowlarr.yml deve ser colocado em: <Prowlarr_Config>/Definitions/Custom/prowlarr.yml"],
    ]
    
    for info_row in prowlarr_info:
        row = add_data_row(ws, row, info_row)
    
    row += 2
    
    # Estrutura do prowlarr.yml
    row = add_subtitle(ws, "Estrutura do arquivo prowlarr.yml", row, cols=4)
    headers = ["Seção", "Campo", "Valor/Descrição", "Observações"]
    row = add_header_row(ws, headers, start_row=row)
    
    estrutura_data = [
        ["Metadados", "id", "dfindexer", "Identificador único do indexador no Prowlarr"],
        ["Metadados", "name", "DF Indexer", "Nome exibido no Prowlarr"],
        ["Metadados", "description", "Python Torrent Indexing for Brazilian", "Descrição do indexador"],
        ["Metadados", "language", "pt-BR", "Idioma do indexador"],
        ["Metadados", "type", "public", "Tipo: public (público) ou private (privado)"],
        ["Metadados", "encoding", "UTF-8", "Codificação de caracteres"],
        ["Metadados", "links", "http://dfindexer:7006/", "URL base da API (ajustar conforme ambiente)"],
        ["Capabilities", "categories.Movies", "Movies", "Categoria para filmes"],
        ["Capabilities", "categories.TV", "TV", "Categoria para séries"],
        ["Capabilities", "modes.search", "[q]", "Modo de busca geral (usa parâmetro q)"],
        ["Capabilities", "modes.tv-search", "[q, season]", "Modo de busca de séries (usa q e season)"],
        ["Capabilities", "modes.movie-search", "[q]", "Modo de busca de filmes (usa parâmetro q)"],
        ["Settings", "scraper_type", "select (1-7)", "Permite escolher qual scraper usar (1=Starck, 2=Rede, etc.)"],
        ["Settings", "use_flaresolverr", "select (true/false)", "Permite habilitar FlareSolverr para sites com Cloudflare"],
        ["Search", "paths.path", "/indexers/{{ .Config.scraper_type }}", "Endpoint da API (usa variável scraper_type)"],
        ["Search", "paths.response.type", "json", "Tipo de resposta esperado (JSON)"],
        ["Search", "inputs.filter_results", "true", "Sempre filtra resultados com similaridade zero"],
        ["Search", "inputs.use_flaresolverr", "{{ .Config.use_flaresolverr }}", "Usa valor da configuração do usuário"],
        ["Search", "inputs.q", "{{ .Keywords }}", "Query de busca (palavras-chave do Prowlarr)"],
        ["Search", "rows.selector", "$.results", "Seletor JSONPath para array de resultados"],
        ["Search", "rows.count.selector", "$.count", "Seletor JSONPath para contagem de resultados"],
    ]
    
    for estrutura_row in estrutura_data:
        row = add_data_row(ws, row, estrutura_row)
    
    # Ajusta largura das colunas
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 50
    
    row += 2
    
    # Fluxo de Integração
    row = add_subtitle(ws, "Fluxo de Integração Prowlarr → API", row, cols=2)
    headers = ["Etapa", "Descrição"]
    row = add_header_row(ws, headers, start_row=row)
    
    fluxo_data = [
        ["1", "Usuário configura indexador no Prowlarr"],
        ["2", "Prowlarr carrega prowlarr.yml da pasta Definitions/Custom/"],
        ["3", "Usuário seleciona scraper (1-7) e opções (FlareSolverr) nas configurações"],
        ["4", "Sonarr/Radarr solicita busca ao Prowlarr (ex: 'Pluribus temporada 1')"],
        ["5", "Prowlarr aplica filtros de palavras-chave (tolower, re_replace)"],
        ["6", "Prowlarr monta URL: GET /indexers/{scraper_type}?q={keywords}&filter_results=true&use_flaresolverr={config}"],
        ["7", "API recebe requisição e processa via IndexerService"],
        ["8", "IndexerService valida scraper_type e cria instância do scraper"],
        ["9", "Scraper busca torrents no site e enriquece com metadata/trackers"],
        ["10", "API retorna JSON: {results: [...], count: N}"],
        ["11", "Prowlarr extrai campos usando seletores JSONPath"],
        ["12", "Prowlarr aplica filtros e transformações nos campos"],
        ["13", "Prowlarr categoriza resultados (Movies/TV) baseado em regex"],
        ["14", "Prowlarr retorna resultados para Sonarr/Radarr"],
    ]
    
    for fluxo_row in fluxo_data:
        row = add_data_row(ws, row, fluxo_row)
    
    row += 2
    
    # Mapeamento de Campos
    row = add_subtitle(ws, "Mapeamento de Campos: API → Prowlarr", row, cols=4)
    headers = ["Campo Prowlarr", "Campo API (JSON)", "Filtros Aplicados", "Observações"]
    row = add_header_row(ws, headers, start_row=row)
    
    campos_data = [
        ["download", "magnet_link", "Nenhum", "Link magnet usado para download"],
        ["title", "title", "Nenhum", "Título padronizado do torrent"],
        ["description", "original_title", "Nenhum", "Título original extraído do HTML"],
        ["details", "details", "Nenhum", "URL da página do torrent"],
        ["infohash", "info_hash", "Nenhum", "Hash SHA-1 do torrent (40 caracteres hex)"],
        ["date", "date", "Nenhum", "Data de criação/publicação do torrent"],
        ["size", "size", "re_replace: '^(|0 B|0B)$' → '2,5 GB'", "Tamanho em bytes. Se vazio/null, usa 2,5 GB como fallback"],
        ["seeders", "seed_count", "re_replace: '^(|null|0)$' → '1'", "Número de seeders. Se vazio/null/0, usa 1 como fallback"],
        ["leechers", "leech_count", "re_replace: '^(|null|0)$' → '1'", "Número de leechers. Se vazio/null/0, usa 1 como fallback"],
        ["imdb", "imdb", "Nenhum", "ID IMDB (formato: tt1234567)"],
        ["category_is_tv_show", "title", "regexp: detecta temporada/episódio", "Detecta se é série baseado em padrões no título"],
        ["category", "Calculado", "Se category_is_tv_show → 'TV', senão → 'Movies'", "Categoria final usada pelo Prowlarr"],
    ]
    
    for campo_row in campos_data:
        row = add_data_row(ws, row, campo_row)
    
    # Ajusta largura das colunas
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 60
    
    row += 2
    
    # Filtros de Palavras-chave
    row = add_subtitle(ws, "Filtros de Palavras-chave (keywordsfilters)", row, cols=3)
    headers = ["Ordem", "Filtro", "Ação"]
    row = add_header_row(ws, headers, start_row=row)
    
    filtros_data = [
        ["1", "tolower", "Converte todas as palavras-chave para minúsculas\nExemplo: 'Pluribus Temporada 1' → 'pluribus temporada 1'"],
        ["2", "re_replace", "Padrão: '(?i)(S0)(\\d{1,2})$'\nSubstitui: 'S01' → 'temporada 1'\nExemplo: 'pluribus S01' → 'pluribus temporada 1'"],
        ["3", "re_replace", "Padrão: '(?i)(S)(\\d{1,3})$'\nSubstitui: 'S1' → 'temporada 1'\nExemplo: 'pluribus S1' → 'pluribus temporada 1'"],
    ]
    
    for filtro_row in filtros_data:
        row = add_data_row(ws, row, filtro_row)
    
    # Ajusta largura da coluna de ação
    ws.column_dimensions['C'].width = 80
    
    row += 2
    
    # Detecção de Categoria (TV vs Movies)
    row = add_subtitle(ws, "Detecção de Categoria: TV vs Movies", row, cols=2)
    headers = ["Padrão Regex", "Exemplo de Match"]
    row = add_header_row(ws, headers, start_row=row)
    
    categoria_data = [
        ["temporada|season", "Pluribus Temporada 1, Game of Thrones Season 2"],
        ["S\\d{1,3}E\\d{1,3}", "S01E01, S2E10, S03E05"],
        ["\\bS\\d{1,3}\\b", "S01, S2, S03 (temporada completa)"],
        ["\\d+x\\d+", "1x01, 2x10, 3x05"],
        ["série|series", "Minha Série, The Series"],
        ["episódio|episode|EP\\d+", "Episódio 1, Episode 2, EP01"],
        ["1ª|2ª|3ª|4ª|5ª|6ª|7ª|8ª|9ª|10ª", "1ª Temporada, 2ª Temporada"],
        ["completa", "Temporada Completa"],
    ]
    
    for cat_row in categoria_data:
        row = add_data_row(ws, row, cat_row)
    
    row += 2
    
    # Configurações Disponíveis no Prowlarr
    row = add_subtitle(ws, "Configurações Disponíveis no Prowlarr", row, cols=4)
    headers = ["Configuração", "Tipo", "Valores", "Descrição"]
    row = add_header_row(ws, headers, start_row=row)
    
    config_data = [
        ["scraper_type", "select", "1-7", "Escolhe qual scraper usar:\n1 = Starck Filmes\n2 = Rede Torrent\n3 = Baixa Filmes\n4 = Torrent dos Filmes\n6 = Comando Torrents\n7 = Bludv Filmes"],
        ["use_flaresolverr", "select", "true/false", "Habilita FlareSolverr para resolver Cloudflare.\nRecomendado para sites que usam proteção Cloudflare.\nAumenta tempo de resposta mas permite acesso a sites protegidos."],
    ]
    
    for config_row in config_data:
        row = add_data_row(ws, row, config_row)
    
    # Ajusta largura das colunas
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 80
    
    row += 2
    
    # Como Adicionar Novo Scraper ao prowlarr.yml
    row = add_subtitle(ws, "Como Adicionar Novo Scraper ao prowlarr.yml", row, cols=3)
    headers = ["Passo", "Ação", "Exemplo"]
    row = add_header_row(ws, headers, start_row=row)
    
    adicionar_scraper_data = [
        ["1", "Adicionar novo scraper ao projeto", "Criar scraper/novosite.py seguindo o fluxo da aba SCRAPERS"],
        ["2", "Adicionar ID ao SCRAPER_NUMBER_MAP", "Em api/services/indexer_service.py:\nSCRAPER_NUMBER_MAP[\"8\"] = \"novosite\""],
        ["3", "Adicionar opção no prowlarr.yml", "Em settings.scraper_type.options:\n\"8\": Novo Site"],
        ["4", "Reiniciar serviços", "Reiniciar API e Prowlarr para aplicar mudanças"],
        ["5", "Configurar no Prowlarr", "Acessar configurações do indexador e selecionar scraper \"8\""],
    ]
    
    for adicionar_row in adicionar_scraper_data:
        row = add_data_row(ws, row, adicionar_row)
    
    # Ajusta largura das colunas
    ws.column_dimensions['C'].width = 80
    
    row += 2
    
    # Exemplo de Requisição
    row = add_subtitle(ws, "Exemplo de Requisição Prowlarr → API", row, cols=2)
    headers = ["Componente", "Valor"]
    row = add_header_row(ws, headers, start_row=row)
    
    exemplo_requisicao = [
        ["URL Base", "http://dfindexer:7006"],
        ["Endpoint", "/indexers/1"],
        ["Query Params", "q=pluribus+temporada+1&filter_results=true&use_flaresolverr=false"],
        ["URL Completa", "http://dfindexer:7006/indexers/1?q=pluribus+temporada+1&filter_results=true&use_flaresolverr=false"],
        ["Método", "GET"],
        ["Resposta Esperada", "JSON com {results: [...], count: N}"],
    ]
    
    for exemplo_row in exemplo_requisicao:
        row = add_data_row(ws, row, exemplo_row)
    
    row += 2
    
    # Instalação e Configuração
    row = add_subtitle(ws, "Instalação e Configuração no Prowlarr", row, cols=3)
    headers = ["Passo", "Ação", "Detalhes"]
    row = add_header_row(ws, headers, start_row=row)
    
    instalacao_data = [
        ["1", "Localizar diretório de configuração", "Geralmente em:\n- Linux: ~/.config/Prowlarr/\n- Windows: %APPDATA%\\Prowlarr\\\n- Docker: /config/prowlarr/"],
        ["2", "Criar diretório Custom", "Criar pasta: Definitions/Custom/"],
        ["3", "Copiar prowlarr.yml", "Copiar arquivo prowlarr.yml para Definitions/Custom/"],
        ["4", "Ajustar URL (se necessário)", "Editar campo 'links' se API não estiver em http://dfindexer:7006/"],
        ["5", "Reiniciar Prowlarr", "Reiniciar serviço Prowlarr para carregar nova definição"],
        ["6", "Adicionar Indexador", "Ir em Settings → Indexers → Add Indexer → Custom"],
        ["7", "Selecionar 'DF Indexer'", "Escolher 'DF Indexer' na lista de indexadores customizados"],
        ["8", "Configurar Scraper", "Selecionar scraper desejado (1-7) nas configurações"],
        ["9", "Configurar FlareSolverr", "Se necessário, habilitar FlareSolverr para sites com Cloudflare"],
        ["10", "Testar Conexão", "Clicar em 'Test' para verificar se conexão funciona"],
    ]
    
    for inst_row in instalacao_data:
        row = add_data_row(ws, row, inst_row)
    
    # Ajusta largura das colunas
    ws.column_dimensions['C'].width = 80
    
    row += 2
    
    # Troubleshooting
    row = add_subtitle(ws, "Troubleshooting - Problemas Comuns", row, cols=3)
    headers = ["Problema", "Causa Provável", "Solução"]
    row = add_header_row(ws, headers, start_row=row)
    
    troubleshooting_data = [
        ["Indexador não aparece", "prowlarr.yml não está em Definitions/Custom/", "Verificar localização do arquivo e reiniciar Prowlarr"],
        ["Erro de conexão", "URL incorreta ou API não está rodando", "Verificar URL em 'links' e status da API"],
        ["Nenhum resultado retornado", "Scraper selecionado não funciona ou site está offline", "Testar scraper diretamente na API, verificar logs"],
        ["Resultados vazios (size=0)", "Filtro muito restritivo ou site sem resultados", "Verificar query, testar sem filter_results"],
        ["Cloudflare bloqueando", "Site usa proteção Cloudflare", "Habilitar use_flaresolverr=true e configurar FLARESOLVERR_ADDRESS"],
        ["Categoria incorreta (TV/Movies)", "Regex de detecção não está funcionando", "Verificar padrões no título, ajustar regex se necessário"],
    ]
    
    for trouble_row in troubleshooting_data:
        row = add_data_row(ws, row, trouble_row)
    
    # Ajusta largura das colunas
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 80
    
    row += 2
    
    # Fluxograma Visual Completo
    row = add_subtitle(ws, "Fluxograma Visual: Integração Prowlarr → API → Scraper", row, cols=1)
    
    fluxograma_prowlarr = """INÍCIO: Usuário busca conteúdo no Sonarr/Radarr
│
├─ Sonarr/Radarr solicita busca ao Prowlarr
│  └─ Exemplo: "Pluribus Temporada 1"
│
├─ Prowlarr processa busca
│  │
│  ├─ Aplica filtros de palavras-chave:
│  │  ├─ tolower: "Pluribus Temporada 1" → "pluribus temporada 1"
│  │  ├─ re_replace: "S01" → "temporada 1"
│  │  └─ re_replace: "S1" → "temporada 1"
│  │
│  ├─ Monta URL da API:
│  │  └─ GET /indexers/{scraper_type}?q={keywords}&filter_results=true&use_flaresolverr={config}
│  │     Exemplo: GET /indexers/1?q=pluribus+temporada+1&filter_results=true&use_flaresolverr=false
│  │
│  └─ Faz requisição HTTP à API
│
├─ API recebe requisição (api/handlers.py)
│  │
│  ├─ Extrai parâmetros:
│  │  ├─ scraper_type: "1" (ou nome do scraper)
│  │  ├─ query: "pluribus temporada 1"
│  │  ├─ filter_results: true
│  │  └─ use_flaresolverr: false
│  │
│  └─ Chama IndexerService.search()
│
├─ IndexerService (api/services/indexer_service.py)
│  │
│  ├─ Valida scraper_type
│  │  └─ Converte "1" → "starck" via SCRAPER_NUMBER_MAP
│  │
│  ├─ Cria instância do scraper
│  │  └─ scraper = create_scraper("starck", use_flaresolverr=False)
│  │
│  └─ Chama scraper.search(query, filter_func)
│
├─ Scraper (scraper/starck.py)
│  │
│  ├─ Busca HTML do site
│  │  └─ Usa cache Redis se disponível
│  │
│  ├─ Extrai links dos torrents
│  │  └─ _extract_links_from_page()
│  │
│  ├─ Para cada link, extrai dados
│  │  └─ _parse_torrent_page()
│  │
│  └─ Enriquece torrents (TorrentEnricher)
│     ├─ Remove duplicados
│     ├─ Busca metadata (size, date) via iTorrents.org
│     ├─ Busca trackers (seeds/leechers) via UDP
│     └─ Aplica filtro de similaridade
│
├─ API retorna JSON
│  │
│  └─ {
│       "results": [
│         {
│           "title": "Pluribus.S01E01.2025.1080p.WEB-DL",
│           "magnet_link": "magnet:?xt=urn:btih:...",
│           "seed_count": 150,
│           "leech_count": 25,
│           ...
│         }
│       ],
│       "count": 1
│     }
│
├─ Prowlarr processa resposta
│  │
│  ├─ Extrai array de resultados
│  │  └─ $.results (JSONPath)
│  │
│  ├─ Mapeia campos:
│  │  ├─ download ← magnet_link
│  │  ├─ title ← title
│  │  ├─ seeders ← seed_count (com fallback)
│  │  ├─ leechers ← leech_count (com fallback)
│  │  └─ ...
│  │
│  ├─ Detecta categoria (TV/Movies)
│  │  └─ Regex no título: "S01E01" → TV
│  │
│  └─ Aplica filtros e transformações
│
└─ Prowlarr retorna resultados para Sonarr/Radarr
   │
   └─ Sonarr/Radarr exibe resultados para download
   
FIM: Usuário pode baixar torrent via magnet link"""
    
    for line in fluxograma_prowlarr.split('\n'):
        row = add_tree_text(ws, row, line)
        row += 1
    
    auto_adjust_columns(ws)
    apply_default_font(ws)
    return ws

def create_resultados_sheet(wb):
    """Cria aba RESULTADOS - Estrutura e Formato dos Resultados"""
    ws = wb.create_sheet("RESULTADOS")
    
    row = add_title(ws, "Estrutura e Formato dos Resultados da API", cols=4)
    
    # Estrutura do JSON de Resposta
    row = add_subtitle(ws, "Estrutura do JSON de Resposta", row, cols=2)
    headers = ["Campo", "Descrição"]
    row = add_header_row(ws, headers, start_row=row)
    
    estrutura_json = [
        ["results", "Array de objetos torrent com os resultados da busca. Cada objeto contém informações de um torrent encontrado."],
        ["count", "Número inteiro indicando a quantidade total de resultados retornados. Corresponde ao tamanho do array 'results'."],
        ["teste", "Campo opcional (boolean). Presente apenas quando a busca foi feita sem query (query vazia). Indica que é uma busca de teste do Prowlarr."],
        ["error", "Campo opcional (string). Presente apenas em caso de erro. Contém a mensagem de erro descritiva."],
    ]
    
    for estrutura_row in estrutura_json:
        row = add_data_row(ws, row, estrutura_row)
    
    row += 2
    
    # Campos do Torrent
    row = add_subtitle(ws, "Campos do Objeto Torrent", row, cols=4)
    headers = ["Campo", "Tipo", "Obrigatório", "Descrição"]
    row = add_header_row(ws, headers, start_row=row)
    
    campos_torrent = [
        ["title", "string", "✅ Sim", "Título padronizado do torrent no formato: base_title.SxxExx.ano.qualidade.codec.áudio. Exemplo: 'Pluribus.S01E01.2025.1080p.WEB-DL.x264.LEGENDADO'"],
        ["magnet_link", "string", "✅ Sim", "Link magnet completo para download do torrent. Formato: 'magnet:?xt=urn:btih:...&dn=...&tr=...'"],
        ["info_hash", "string", "✅ Sim", "Hash SHA-1 do torrent em hexadecimal (40 caracteres). Exemplo: 'a1b2c3d4e5f6...'"],
        ["details", "string", "✅ Sim", "URL da página do torrent no site original. Link para visualizar detalhes no site de origem."],
        ["original_title", "string", "⚠️ Opcional", "Título original extraído do HTML da página. Pode estar em português ou outro idioma. Vazio se não encontrado."],
        ["size", "string", "⚠️ Opcional", "Tamanho do torrent formatado (ex: '2.45 GB', '1.2 TB'). Vazio se não encontrado. Formato legível para humanos."],
        ["date", "string", "⚠️ Opcional", "Data de criação/publicação do torrent no formato ISO 8601. Exemplo: '2025-07-10T18:30:00'. Vazio se não encontrado."],
        ["seed_count", "integer", "⚠️ Opcional", "Número de seeders (pessoas compartilhando o arquivo completo). Valor padrão: 0 se não encontrado."],
        ["leech_count", "integer", "⚠️ Opcional", "Número de leechers (pessoas baixando o arquivo). Valor padrão: 0 se não encontrado."],
        ["imdb", "string", "⚠️ Opcional", "ID do IMDB no formato 'tt1234567'. Vazio se não encontrado. Usado para identificação de filmes/séries."],
        ["year", "string", "⚠️ Opcional", "Ano de lançamento do conteúdo (4 dígitos). Exemplo: '2025'. Vazio se não encontrado."],
    ]
    
    for campo_row in campos_torrent:
        row = add_data_row(ws, row, campo_row)
    
    # Ajusta largura das colunas
    ws.column_dimensions['D'].width = 80
    
    row += 2
    
    # Tipos de Conteúdo e Idiomas
    row = add_subtitle(ws, "Tipos de Conteúdo e Idiomas", row, cols=3)
    headers = ["Tipo", "Identificação", "Exemplo no Título"]
    row = add_header_row(ws, headers, start_row=row)
    
    tipos_conteudo = [
        ["Série (TV Show)", "Presença de padrão SxxExx ou Sxx no título", "Pluribus.S01E01.2025.1080p.WEB-DL"],
        ["Série Completa", "Presença de Sxx sem Exx (temporada completa)", "Pluribus.S02.2025.1080p.WEB-DL"],
        ["Filme", "Ausência de padrão SxxExx, presença de ano", "Matrix.1999.1080p.BluRay"],
        ["LEGENDADO", "Tag [LEGENDADO] ou .LEGENDADO no título", "Filme.2025.1080p.WEB-DL.LEGENDADO"],
        ["DUBLADO", "Tag [DUBLADO] ou .DUBLADO no título", "Filme.2025.1080p.WEB-DL.DUBLADO"],
        ["DUAL", "Tag [DUAL] ou .DUAL no título (áudio duplo)", "Filme.2025.1080p.WEB-DL.DUAL"],
        ["NACIONAL", "Tag [NACIONAL] no título", "Filme.2025.1080p.WEB-DL.NACIONAL"],
        ["[Brazilian]", "Tag [Brazilian] no título (português brasileiro)", "Filme.2025.1080p.WEB-DL.[Brazilian]"],
        ["[Eng]", "Tag [Eng] no título (inglês)", "Filme.2025.1080p.WEB-DL.[Eng]"],
        ["[br-dub]", "Tag [br-dub] no título (dublado brasileiro)", "Filme.2025.1080p.WEB-DL.[br-dub]"],
    ]
    
    for tipo_row in tipos_conteudo:
        row = add_data_row(ws, row, tipo_row)
    
    row += 2
    
    # Tags Adicionadas Automaticamente
    row = add_subtitle(ws, "Tags Adicionadas Automaticamente", row, cols=3)
    headers = ["Tag", "Quando é Adicionada", "Exemplo"]
    row = add_header_row(ws, headers, start_row=row)
    
    tags_adicionadas = [
        ["[Brazilian]", "Adicionada quando detecta DUAL, DUBLADO ou NACIONAL no release_title_magnet ou metadata", "Filme.2025.1080p.WEB-DL.DUBLADO [Brazilian]"],
        ["[Eng]", "Adicionada quando detecta LEGENDADO, LEGENDA ou LEG no release_title_magnet ou metadata", "Filme.2025.1080p.WEB-DL.LEGENDADO [Eng]"],
        ["[Brazilian] [Eng]", "Adicionada quando detecta palavras dos 2 tipos acima (ex: DUAL = português + inglês)", "Filme.2025.1080p.WEB-DL.DUAL [Brazilian] [Eng]"],
    ]
    
    for tag_row in tags_adicionadas:
        row = add_data_row(ws, row, tag_row)
    
    # Ajusta largura das colunas
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 60
    
    row += 2
    
    # Como as Tags São Escolhidas
    row = add_subtitle(ws, "Como as Tags São Escolhidas", row, cols=2)
    headers = ["Etapa", "Descrição"]
    row = add_header_row(ws, headers, start_row=row)
    
    como_escolhidas = [
        ["1. Detecção no release_title_magnet", "Sistema verifica primeiro o nome do arquivo no magnet link (display_name) para detectar palavras: DUAL, DUBLADO, NACIONAL, LEGENDADO, LEGENDA, LEG"],
        ["2. Busca no metadata (se não encontrou)", "Se não encontrou no release_title_magnet, busca no metadata do torrent via iTorrents.org usando o info_hash"],
        ["3. Adição das tags", "Adiciona [Brazilian] se detectar DUAL/DUBLADO/NACIONAL. Adiciona [Eng] se detectar LEGENDADO/LEGENDA/LEG. Se detectar DUAL, adiciona ambas as tags"],
    ]
    
    for escolha_row in como_escolhidas:
        row = add_data_row(ws, row, escolha_row)
    
    # Ajusta largura das colunas
    ws.column_dimensions['B'].width = 100
    
    row += 2
    
    # Exemplo Completo de Resposta JSON
    row = add_subtitle(ws, "Exemplo Completo de Resposta JSON", row, cols=1)
    
    exemplo_json = """{
  "results": [
    {
      "title": "Pluribus.S01E01.2025.1080p.WEB-DL.x264.LEGENDADO [Eng]",
      "magnet_link": "magnet:?xt=urn:btih:a1b2c3d4e5f6...&dn=Pluribus+S01E01&tr=udp://tracker1:80&tr=udp://tracker2:80",
      "info_hash": "a1b2c3d4e5f6789012345678901234567890abcd",
      "details": "https://starckfilmes-v3.com/torrent/pluribus-s01e01",
      "original_title": "Pluribus",
      "size": "2.45 GB",
      "date": "2025-07-10T18:30:00",
      "seed_count": 150,
      "leech_count": 25,
      "imdb": "tt1234567",
      "year": "2025"
    },
    {
      "title": "Matrix.1999.1080p.BluRay.x265.DUBLADO [Brazilian]",
      "magnet_link": "magnet:?xt=urn:btih:b2c3d4e5f6a1...&dn=Matrix+1999&tr=udp://tracker1:80",
      "info_hash": "b2c3d4e5f6a1789012345678901234567890bcde",
      "details": "https://starckfilmes-v3.com/torrent/matrix-1999",
      "original_title": "Matrix",
      "size": "8.5 GB",
      "date": "2025-07-09T14:20:00",
      "seed_count": 320,
      "leech_count": 45,
      "imdb": "tt0133093",
      "year": "1999"
    },
    {
      "title": "Filme.2025.1080p.WEB-DL.DUAL [Brazilian] [Eng]",
      "magnet_link": "magnet:?xt=urn:btih:c3d4e5f6a1b2...&dn=Filme+2025&tr=udp://tracker1:80",
      "info_hash": "c3d4e5f6a1b2789012345678901234567890cdef",
      "details": "https://starckfilmes-v3.com/torrent/filme-2025",
      "original_title": "Filme",
      "size": "3.2 GB",
      "date": "2025-07-08T10:15:00",
      "seed_count": 200,
      "leech_count": 30,
      "year": "2025"
    }
  ],
  "count": 3
}"""
    
    exemplo_cell = ws.cell(row=row, column=1, value=exemplo_json)
    exemplo_cell.font = Font(name='Courier New', size=8)
    exemplo_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.merge_cells(f'A{row}:A{row + len(exemplo_json.split(chr(10))) - 1}')
    row += len(exemplo_json.split(chr(10))) + 2
    
    # Exemplo de Resposta com Query Vazia (Teste)
    row = add_subtitle(ws, "Exemplo de Resposta com Query Vazia (Teste Prowlarr)", row, cols=1)
    
    exemplo_teste = """{
  "results": [
    {
      "title": "Filme.2025.1080p.WEB-DL",
      "magnet_link": "magnet:?xt=urn:btih:...",
      "info_hash": "...",
      "details": "https://...",
      "original_title": "Filme",
      "size": "2.5 GB",
      "date": "2025-07-10T12:00:00",
      "seed_count": 50,
      "leech_count": 10
    }
  ],
  "count": 1,
  "teste": true
}"""
    
    exemplo_teste_cell = ws.cell(row=row, column=1, value=exemplo_teste)
    exemplo_teste_cell.font = Font(name='Courier New', size=8)
    exemplo_teste_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.merge_cells(f'A{row}:A{row + len(exemplo_teste.split(chr(10))) - 1}')
    row += len(exemplo_teste.split(chr(10))) + 2
    
    # Exemplo de Resposta com Erro
    row = add_subtitle(ws, "Exemplo de Resposta com Erro", row, cols=1)
    
    exemplo_erro = """{
  "error": "Scraper 'invalido' não configurado. Tipos disponíveis: ['starck', 'rede', 'baixafilmes', 'tfilme', 'comand', 'bludv']",
  "results": [],
  "count": 0
}"""
    
    exemplo_erro_cell = ws.cell(row=row, column=1, value=exemplo_erro)
    exemplo_erro_cell.font = Font(name='Courier New', size=8)
    exemplo_erro_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.merge_cells(f'A{row}:A{row + len(exemplo_erro.split(chr(10))) - 1}')
    row += len(exemplo_erro.split(chr(10))) + 2
    
    # Formato dos Campos
    row = add_subtitle(ws, "Formato e Validação dos Campos", row, cols=3)
    headers = ["Campo", "Formato", "Valores Válidos"]
    row = add_header_row(ws, headers, start_row=row)
    
    formatos_campos = [
        ["title", "String padronizada", "Formato: base_title.SxxExx.ano.qualidade.codec.áudio\nExemplo: 'Pluribus.S01E01.2025.1080p.WEB-DL.x264.LEGENDADO'"],
        ["magnet_link", "URI Magnet", "Deve começar com 'magnet:?' e conter 'xt=urn:btih:' obrigatoriamente"],
        ["info_hash", "Hexadecimal (40 chars)", "Exatamente 40 caracteres hexadecimais (0-9, a-f). Exemplo: 'a1b2c3d4e5f6789012345678901234567890abcd'"],
        ["details", "URL completa", "URL válida começando com 'http://' ou 'https://'"],
        ["size", "String formatada", "Formato legível: número + unidade (ex: '2.45 GB', '1.2 TB', '500 MB')"],
        ["date", "ISO 8601", "Formato: 'YYYY-MM-DDTHH:MM:SS' ou 'YYYY-MM-DD'. Exemplo: '2025-07-10T18:30:00'"],
        ["seed_count", "Integer >= 0", "Número inteiro não negativo. Padrão: 0 se não encontrado"],
        ["leech_count", "Integer >= 0", "Número inteiro não negativo. Padrão: 0 se não encontrado"],
        ["imdb", "ID IMDB", "Formato: 'tt' seguido de 7-8 dígitos. Exemplo: 'tt1234567'. Vazio se não encontrado"],
        ["year", "Ano (4 dígitos)", "Ano de 4 dígitos (1900-2100). Exemplo: '2025'. Vazio se não encontrado"],
    ]
    
    for formato_row in formatos_campos:
        row = add_data_row(ws, row, formato_row)
    
    # Ajusta largura das colunas
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 80
    
    row += 2
    
    # Padrões de Título
    row = add_subtitle(ws, "Padrões de Título por Tipo de Conteúdo", row, cols=3)
    headers = ["Tipo", "Padrão", "Exemplo"]
    row = add_header_row(ws, headers, start_row=row)
    
    padroes_titulo = [
        ["Episódio Simples", "base_title.SxxExx.ano.qualidade.codec.áudio", "Pluribus.S01E01.2025.1080p.WEB-DL.LEGENDADO"],
        ["Episódios Múltiplos (2)", "base_title.SxxExx-Exx.ano.qualidade.codec", "Pluribus.S02E05-E06.2025.1080p.WEB-DL"],
        ["Episódios Múltiplos (3+)", "base_title.SxxExxEExxEExx.ano.qualidade", "Pluribus.S02E05E06E07.2025.1080p.WEB-DL"],
        ["Temporada Completa", "base_title.Sxx.ano.qualidade.codec", "Pluribus.S02.2025.1080p.WEB-DL"],
        ["Filme", "base_title.ano.qualidade.codec.áudio", "Matrix.1999.1080p.BluRay.DUBLADO"],
        ["Com Tags de Idioma", "base_title.ano.qualidade.[tag]", "Filme.2025.1080p.WEB-DL.[Brazilian]"],
    ]
    
    for padrao_row in padroes_titulo:
        row = add_data_row(ws, row, padrao_row)
    
    row += 2
    
    # Campos Opcionais - Quando Aparecem
    row = add_subtitle(ws, "Campos Opcionais - Quando Aparecem", row, cols=3)
    headers = ["Campo", "Condição", "Observação"]
    row = add_header_row(ws, headers, start_row=row)
    
    campos_opcionais = [
        ["original_title", "Se encontrado no HTML", "Extraído diretamente do HTML da página pelo scraper. Pode estar em português ou outro idioma."],
        ["size", "Se encontrado via metadata, magnet ou HTML", "Prioridade: 1) Metadata API, 2) Parâmetro 'xl' do magnet, 3) HTML do site"],
        ["date", "Se encontrado via metadata ou HTML", "Prioridade: 1) Metadata API (creation_date), 2) HTML do site"],
        ["seed_count", "Se encontrado via tracker scraping", "Obtido via consulta UDP aos trackers. Valor padrão: 0 se não encontrado"],
        ["leech_count", "Se encontrado via tracker scraping", "Obtido via consulta UDP aos trackers. Valor padrão: 0 se não encontrado"],
        ["imdb", "Se encontrado no HTML, cache ou metadata", "Prioridade: 1) HTML, 2) Cache por info_hash, 3) Cache por título, 4) Metadata"],
        ["year", "Se extraído do título ou HTML", "Extraído do título padronizado ou do HTML da página"],
    ]
    
    for campo_row in campos_opcionais:
        row = add_data_row(ws, row, campo_row)
    
    # Ajusta largura das colunas
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 80
    
    auto_adjust_columns(ws)
    apply_default_font(ws)
    return ws

def main():
    """Função principal"""
    print("Criando SUPER README detalhado em XLSX...")
    
    wb = create_workbook()
    
    # Cria todas as abas
    create_caracteristicas_sheet(wb)
    create_redis_sheet(wb)
    create_tracker_sheet(wb)
    create_title_sheet(wb)
    create_definicoes_sheet(wb)
    create_scrapers_sheet(wb)
    create_prowlarr_sheet(wb)
    create_resultados_sheet(wb)
    
    # Salva arquivo
    filename = "Projeto_DFlexy.xlsx"
    wb.save(filename)
    print(f"✅ Documentação criada: {filename}")
    print(f"   Total de abas: {len(wb.sheetnames)}")
    print(f"   Abas criadas: {', '.join(wb.sheetnames)}")

if __name__ == "__main__":
    main()

